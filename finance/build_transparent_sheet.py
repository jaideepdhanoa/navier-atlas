#!/usr/bin/env python3
"""
Build a fully transparent, formula-driven unit-economics workbook for a partner.
Reads ALL numbers from the model files (no hardcoded data) and emits an .xlsx
with live cell formulas + named ranges + a scenario toggle, so every output is
traceable back to a named assumption. Upload-convert to a native Google Sheet.

Usage: uv run --with openpyxl python3 build_transparent_sheet.py --partner grab --out /tmp/grab_unit_econ.xlsx
"""
import json, math, os, sys

HERE = os.path.dirname(os.path.abspath(__file__))
MODEL = os.path.join(HERE, "model")

def arg(flag, default=None):
    return sys.argv[sys.argv.index(flag)+1] if flag in sys.argv else default

PARTNER = arg("--partner", "grab")
OUT = arg("--out", f"/tmp/{PARTNER}_unit_econ.xlsx")
GLOBAL_MODE = PARTNER == "global"
# LB-260 (Jaideep 2026-06-24): hospitality/captive-luxury proposals use the $1M N30 list-price CAPEX,
# region-independent. Pass --capex-tier hospitality for hospitality partners (FP, Ocean Whisperer, Minor, …)
# so this engine recomputes payback against $1M, matching model/aggregate.py capex_for() (golden rule #7).
HOSPITALITY_CAPEX = arg("--capex-tier", "") == "hospitality"
DEDUP_MODE = arg("--dedup", "unique" if GLOBAL_MODE else "none")
SKIP_README = "--skip-readme" in sys.argv
PARTNER_LABEL = "Global (unique geometry)" if (GLOBAL_MODE and DEDUP_MODE == "unique") else (
    "Global" if GLOBAL_MODE else PARTNER.title())
# --corridors <path>: override the corridor registry source. Mirrors aggregate.py so the
# sheet builds from a SCOPED VIEW of the shared global network for inheriting partners
# (Uber/hotels) WITHOUT duplicating corridors into the durable corridors.json. Default = canonical.
CORR_PATH = arg("--corridors", os.path.join(MODEL, "corridors.json"))

const = json.load(open(os.path.join(MODEL, "vessel-constants.json")))
corr  = json.load(open(CORR_PATH))
cref  = json.load(open(os.path.join(MODEL, "country-reference.json")))["countries"]
gcfg  = json.load(open(os.path.join(MODEL, "growth-config.json")))
COUNTRY_REQUIRED_FIELDS = ("captain_usd_yr", "energy_usd_kwh", "grid_co2_kg_kwh", "marina_overhead_usd_yr", "cost_index")

def validate_country(country):
    if not country or country not in cref:
        raise ValueError(f"Missing exact country-reference row for {country!r}; no fallback is permitted")
    row = cref[country]
    for key in COUNTRY_REQUIRED_FIELDS:
        node = row.get(key)
        value = node.get("value") if isinstance(node, dict) else node
        if not isinstance(value, (int, float)):
            raise ValueError(f"Incomplete country-reference value: {country}.{key}; hold the corridor or source the field")

def v(node):    return node["value"] if isinstance(node, dict) and "value" in node else node
def tier(node): return node.get("source_tier","") if isinstance(node, dict) else ""
def conf(node): return node.get("confidence","")  if isinstance(node, dict) else ""
def src(node):  return node.get("source","")      if isinstance(node, dict) else ""

P2 = const["vessels"]["pioneer_ii"]; OPS = const["operating_defaults"]; CAR = const["carbon"]
SCEN = OPS["_utilization_scenarios"]

pax_cap=v(P2["pax_capacity"]); range_nm=v(P2["range_nm"]); cruise_kt=v(P2["cruise_speed_kt"])
capex=v(P2["capex_usd"]); battery=v(P2["battery_kwh"]); maint=v(P2["annual_maintenance_usd"])
if HOSPITALITY_CAPEX: capex=1000000   # LB-260: hospitality N30 list-price headline
dep_years=v(P2["depreciation_years"]); diesel_nmpg=v(P2["diesel_comparable_nm_per_gal"])
service_hr=v(OPS["service_window_hr_per_day"]); turn_min=v(OPS["turnaround_charge_min"])
mech_uptime=v(OPS["monthly_operational_capacity"]); capture=v(OPS["navier_capture_rate"])
discount=v(OPS["discount_factor"]["ridehail"]); diesel_co2pg=v(CAR["diesel_kg_co2_per_gal"])
# 2026-06-19 calibration: boarding dwell + sailings cap + standardized opex lines
dwell_min=v(OPS["boarding_dwell_min"]); max_tpd=v(OPS["max_trips_per_day"])
# LB-256 archetype-keyed sailings cap + LB-255 premium re-fare (mirror atom.py)
TPD_MAP=OPS.get("max_trips_per_day_by_archetype") or {}
REFARE=OPS.get("premium_refare") or {}
REFARE_ON=bool(REFARE.get("enabled"))
REFARE_CEIL=v(REFARE["subsidized_fare_ceiling"]) if REFARE_ON else None
REFARE_FLOOR=v(REFARE["premium_ondemand_floor"]) if REFARE_ON else None
crew_factor=v(OPS["crew_fte_factor"]); ins_pct=v(OPS["insurance_pct_of_capex"])
charge_berth=v(OPS["charging_berth_annual_usd"])

bands = {b: (SCEN[b]["load_factor"], SCEN[b]["revenue_leg_factor"]) for b in ("thin","mid","full")}
COUNTRY_OPS = OPS.get("country_operating_overrides") or {}
KOREA_COUNTRY = "South Korea"
KOREA_OPS = COUNTRY_OPS.get(KOREA_COUNTRY) or {}
if KOREA_OPS:
    korea_mid_load = v(KOREA_OPS["mid_load_factor"])
    korea_mid_revleg = v(KOREA_OPS["mid_revenue_leg_factor"])
    korea_service_hr = v(KOREA_OPS["service_window_hr_per_day"])
    korea_turn_min = v(KOREA_OPS["turnaround_min"])
    korea_dwell_min = v(KOREA_OPS["boarding_dwell_min"])
    korea_charge_cfg = KOREA_OPS["charge_recovery"]
    korea_charge_range_nm = v(korea_charge_cfg["range_nm"])
    korea_full_range_charge_min = v(korea_charge_cfg["full_range_charge_min"])
else:
    # Non-SK sheet builds still import this module; only Korea corridors use these.
    # Missing override → global defaults (null beats inventing SK-only params).
    korea_mid_load = v(SCEN["mid"]["load_factor"])
    korea_mid_revleg = v(SCEN["mid"]["revenue_leg_factor"])
    korea_service_hr = service_hr
    korea_turn_min = turn_min
    korea_dwell_min = dwell_min
    korea_charge_range_nm = range_nm
    korea_full_range_charge_min = turn_min

# LB-254: effective capture (floor / true transport-spend pool) + captive flag from the partner
# aggregate, so the Market-sizing ladder (the SECOND cost engine — golden rule #7) anchors on the
# TRUE pool exactly like growth.py: M_today = floor / eff_capture = pool, never the inflated
# floor/0.10 that put captive markets ~9x too high. Falls back to the static rate if no agg/pool.
_agg_default = os.path.join(HERE, "recal", "agg-unique-global.json" if (GLOBAL_MODE and DEDUP_MODE == "unique")
                            else f"agg-{PARTNER}.json")
_agg_path = arg("--agg", _agg_default)
EFF_CAPTURE = None; IS_CAPTIVE = False

# ---- market display names (proper-case Market column) ----
MARKET_DISPLAY = {
    "singapore":"Singapore", "cross-border":"Cross-Border", "bali":"Bali",
    "phuket":"Phuket", "philippines":"Philippines", "vietnam":"Vietnam",
    "cambodia":"Cambodia", "borneo":"Borneo", "penang":"Penang", "jakarta":"Jakarta",
    "taiwan":"Taiwan", "saudi-redsea":"Saudi \u2013 Red Sea",
    "saudi-redsea-resort":"Saudi \u2013 Red Sea (Resort)",
    "uae-careem":"UAE (Careem)", "uae-luxury":"UAE (Luxury)",
    "maldives-jih":"Maldives (JIH)",
    "french-polynesia":"French Polynesia",
}
def mkt_disp(mid): return MARKET_DISPLAY.get(mid, mid.replace("-"," ").title())

def pier_key(label: str) -> tuple[str, str]:
    sep = " \u2192 " if " \u2192 " in label else (" -> " if " -> " in label else None)
    if sep:
        a, b = label.split(sep, 1)
        return a.strip().lower(), b.strip().lower()
    return label.strip().lower(), ""

# LB-258: corridor floor-bucket status from aggregate.py — the Market-sizing ladder MUST
# anchor on status=grounded only (same as growth.py _headline_anchor=grounded and deck
# gen_deck_economics.py). Cascade-estimated rows were wrongly summed into the floor.
STATUS_BY_ROUTE: dict[tuple[str, tuple[str, str]], str] = {}
try:
    _agg = json.load(open(_agg_path))
    _gfloor = _agg["rollup"]["grounded_floor"]
    EFF_CAPTURE = _gfloor.get("effective_capture")
    IS_CAPTIVE = bool(EFF_CAPTURE and EFF_CAPTURE >= 0.5)
    for ar in _agg.get("rows", []):
        if ar.get("is_dup"):
            continue
        STATUS_BY_ROUTE[(ar["market"], pier_key(ar.get("corridor", "")))] = ar.get("status", "")
except Exception:
    pass

def floor_bucket(mid, corridor, tier_excl, fwd, demand, status=None):
    """Mirror aggregate.py status buckets for the published floor vs held-out demand."""
    if fwd or status == "forward_sam":
        return "Forward SAM"
    if status == "grounded":
        return "Grounded"
    if status == "estimated":
        return "Upside tier" if tier_excl else "Estimated"
    if tier_excl:
        return "Upside tier"
    return "Grounded" if demand else "Estimated"

# ---- corridors (range gate + dedupe, mirrors aggregate.py) ----
rows=[]; roadmap=[]; economics_holds=[]; seen={}
mkt_fleet_basis={}   # display name -> (fleet_basis, fleet_rounding)  [R-FLOOR-2 / G51]
for mid, mk in corr["markets"].items():
    if not GLOBAL_MODE and mk.get("partner","grab") != PARTNER: continue
    mkt_fleet_basis[mkt_disp(mid)] = (mk.get("fleet_basis","per_corridor_floor"), mk.get("fleet_rounding"))
    for c in mk["corridors"]:
        if c.get("_economics_hold_reason"):
            economics_holds.append({
                "market": mkt_disp(mid), "route_id": c.get("route_id"),
                "corridor": f"{c.get('from','')} → {c.get('to','')}",
                "country": c.get("country"), "reason": c["_economics_hold_reason"],
            })
            continue
        if c.get("_premium_cascade"): continue  # R5-EXT public-transit/no-premium-tier — excluded
        key=(c.get("from","").strip().lower(), c.get("to","").strip().lower())
        if DEDUP_MODE == "unique":
            is_dup = c.get("_dup_of") is not None or (key in seen)
            seen.setdefault(key, mid)
        elif GLOBAL_MODE:
            is_dup = c.get("_dup_of") is not None
        else:
            is_dup = c.get("_dup_of") is not None or (key in seen and seen[key]!=mid)
            seen.setdefault(key, mid)
        if is_dup: continue
        nm=c["distance_nm"]; L3=c.get("L3_locals") or {}
        fr=L3.get("_fare_record") or {}; dr=L3.get("_demand_record") or {}
        country=c.get("country"); validate_country(country)
        w=L3.get("weather_uptime_factor")
        _season_days=L3.get("season_days")  # L3 override wins over 365×uptime×weather (atom.py)
        # demand pool — mirror atom.py LENS 2: average of native proxies (arrivals/ferry)
        # when present; corridor_annual_oneway_pax only as the fallback crossing proxy.
        da=L3.get("demand_arrivals_rides_yr"); df=L3.get("demand_ferry_rides_yr")
        pool_vals=[x for x in (da,df) if x]
        pool=(sum(pool_vals)/len(pool_vals)) if pool_vals else L3.get("corridor_annual_oneway_pax")
        # per-corridor capture — mirror atom.py A' captive override (capture_override_enabled):
        # captive:true (or captive archetype) sole-operator transfer legs capture ~90%, not 10%.
        cap_row=capture
        if OPS.get("capture_override_enabled"):
            if c.get("captive") is True:
                _cc=OPS.get("captive_capture_rate"); cap_row=v(_cc) if _cc is not None else cap_row
            elif c.get("captive") is False:
                pass
            else:
                _ovmap=OPS.get("navier_capture_rate_by_archetype") or {}
                _arche=c.get("archetype")
                if _arche in _ovmap: cap_row=v(_ovmap[_arche])
        # LB-103: pool_basis-aware capture (mirrors atom.py) — sourced pools already
        # narrowed to addressable use capture_on_addressable; pools with capture already
        # applied upstream use 1.0; untagged gross pools keep the 10% default.
        _pb=L3.get("pool_basis")
        if _pb=="capture_applied":
            cap_row=1.0
        elif _pb=="addressable":
            _ca=OPS.get("capture_on_addressable")
            if _ca is not None: cap_row=v(_ca)
        # Per-corridor capture override (mirror atom.py) — e.g. OW L3_locals navier_capture_override 0.55
        _l3cap=L3.get("navier_capture_override")
        if _l3cap is not None:
            cap_row=_l3cap
        # LB-255 premium re-fare: lift subsidized public-transit fares to the premium
        # on-demand floor (mirror atom.py). Genuine premium scheduled fares untouched.
        _fare=L3.get("comparable_fare_usd_pax"); _fare_src=fr.get("source")
        if REFARE_ON and _fare is not None and _fare <= REFARE_CEIL and _fare < REFARE_FLOOR:
            _fare_src=f"[LB-255 PREMIUM RE-FARE: subsidized comparable ${_fare} lifted to premium on-demand floor ${REFARE_FLOOR}] " + (_fare_src or "")
            _fare=REFARE_FLOOR
        # LB-256 archetype-keyed sailings cap (mirror atom.py)
        _arche_cap=c.get("archetype")
        eff_cap=TPD_MAP[_arche_cap] if (_arche_cap in TPD_MAP and not str(_arche_cap).startswith("_")) else max_tpd
        _corridor=f"{c['from']} \u2192 {c['to']}"
        _fwd=("FWD-SAM (2030)" if c.get("_forward_sam") else None)
        _tier_excl=(c.get("_tier") if c.get("_in_grounded_floor") is False else None)
        _status=STATUS_BY_ROUTE.get((mid, pier_key(_corridor)))
        rec=dict(market=mkt_disp(mid), country=country, corridor=_corridor,
                 route_id=c.get("route_id"), nm=nm, fare=_fare, eff_cap=eff_cap,
                 provenance=mk.get("partner", "grab"),
                 fare_tier=fr.get("source_tier"), fare_conf=fr.get("confidence"), fare_src=_fare_src,
                 demand=pool, cap=cap_row,
                 demand_tier=dr.get("source_tier"), demand_conf=dr.get("confidence"), demand_src=dr.get("source"),
                 weather=(w if w is not None else 1.0),
                 season_days=_season_days,
                 subset=c.get("_subset_of"),
                 fwd=_fwd,
                 # LB-99: sourced-greenfield tiers (modal-shift / experience-upside) are
                 # estimated UPSIDE, never grounded floor. experience_grounded stays in.
                 tier_excl=_tier_excl,
                 floor_bucket=floor_bucket(mid, _corridor, _tier_excl, _fwd, pool, _status),
                 # LB-88: captive-resort vessel ceiling = ceil(villas/25)
                 fleet_cap=(math.ceil(c["villas"]/25) if (c.get("captive_resort") and c.get("villas")) else None))
        (roadmap if nm>range_nm else rows).append(rec)
rows.sort(key=lambda r:(r["country"], r["nm"]))

# Enrich global unique rows with geometry-owner metadata from agg-unique-global.json
GEOMETRY_META: dict[tuple[str, str], dict] = {}
if GLOBAL_MODE and DEDUP_MODE == "unique" and os.path.isfile(_agg_path):
    for ar in json.load(open(_agg_path)).get("rows", []):
        if ar.get("is_dup"):
            continue
        pk = pier_key(ar.get("corridor", ""))
        GEOMETRY_META[pk] = {
            "provenance": ar.get("_provenance_partner") or "",
            "alias_markets": ", ".join(ar.get("_alias_markets") or []),
            "alias_count": ar.get("_alias_count") or 1,
        }
    for rec in rows:
        meta = GEOMETRY_META.get(pier_key(rec["corridor"]), {})
        rec["provenance"] = meta.get("provenance") or rec.get("provenance", "")
        rec["alias_markets"] = meta.get("alias_markets", "")
        rec["alias_count"] = meta.get("alias_count", 1)

# country opex (only countries used)
used_countries = sorted({r["country"] for r in rows})

# ============================================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
NAVY="1F3A5F"; STEEL="2E5A88"; LIGHT="DCE6F1"; INPUTB="FFF2CC"; CALC="EAF3EA"; GREY="F2F2F2"; BANDF="FCE4D6"
H1=Font(bold=True,size=16,color="FFFFFF"); H2=Font(bold=True,size=12,color="FFFFFF")
HDR=Font(bold=True,size=9,color="FFFFFF"); BOLD=Font(bold=True); SMALL=Font(size=9,color="555555")
MONO=Font(name="Consolas",size=10)
f_navy=PatternFill("solid",fgColor=NAVY); f_steel=PatternFill("solid",fgColor=STEEL)
f_light=PatternFill("solid",fgColor=LIGHT); f_input=PatternFill("solid",fgColor=INPUTB)
f_calc=PatternFill("solid",fgColor=CALC); f_grey=PatternFill("solid",fgColor=GREY); f_band=PatternFill("solid",fgColor=BANDF)
wrap=Alignment(wrap_text=True,vertical="top"); ctr=Alignment(horizontal="center",vertical="center")
ctrw=Alignment(horizontal="center",vertical="center",wrap_text=True)
side=Side(style="thin",color="BBBBBB"); border=Border(left=side,right=side,top=side,bottom=side)
USD='"$"#,##0'; USD2='"$"#,##0.00'; PCT='0.0%'; NUM='#,##0'; NUM1='#,##0.0'; NUM2='#,##0.00'

def sc(ws, ref, val, font=None, fill=None, fmt=None, align=None, bd=False):
    c=ws[ref]; c.value=val
    if font:c.font=font
    if fill:c.fill=fill
    if fmt:c.number_format=fmt
    if align:c.alignment=align
    if bd:c.border=border
    return c

# Model parity (LB post-#224 / Swing $3 SOM delta): Python round() is banker's
# round-half-even; Excel/Google Sheets ROUND is half-away-from-zero.
# 5 trips/day × 274 op-days × 0.65 rev-leg = 890.5 → model 890, Sheets ROUND → 891.
# Use this helper at every boundary that mirrors Python round(x) / round(x, 0).
def round_half_even(expr: str) -> str:
    """Positive-number round-half-even expression matching Python 3 round()."""
    x = f"({expr})"
    return (
        f'IF(ABS(MOD(ABS({x}),1)-0.5)<1E-9,'
        f'2*ROUND({x}/2,0),'
        f'ROUND({x},0))'
    )

named={}
def addname(name, ref): named[name]=ref

# ------------------------------------------------------------ TAB 0 Read me
if SKIP_README:
    wb.remove(wb.active)
    ws0 = None
else:
    ws0=wb.active; ws0.title="Read me"; ws0.sheet_view.showGridLines=False
    ws0.column_dimensions["A"].width=3; ws0.column_dimensions["B"].width=120
    sc(ws0,"B2",f"Navier \u00d7 {PARTNER_LABEL} \u2014 Corridor Unit Economics",font=H1,fill=f_navy,align=Alignment(vertical="center")); ws0.row_dimensions[2].height=34
content=[
 ("",None,None),
 ("What this is",BOLD,None),
 ("A standalone, transparent unit-economics model for the per-seat electric-shuttle business on each partner corridor. Every output is a live formula you can click and trace back to one named assumption cell. Nothing is hardcoded.",None,None),
 ("",None,None),
 ("The one equation (per boat, per year)",BOLD,None),
 ("REVENUE = gross legs/day \u00d7 operating days/yr \u00d7 revenue-leg utilization \u00d7 (pax capacity \u00d7 seat occupancy) \u00d7 Navier fare",MONO,f_light),
 ("    gross legs/day = FLOOR(service window min \u00f7 cycle time); South Korea cycle = one-way + 20 min turnaround + 10 min boarding + charge recovery",MONO,None),
 ("    South Korea charge recovery = distance_nm / 70 \u00d7 45 min — PLANNING PROXY requiring engineering validation; not a certified charge curve",MONO,None),
 ("    one-way time   = distance(nm) \u00f7 cruise speed(kt)",MONO,None),
 ("    operating days = 365 \u00d7 mechanical uptime \u00d7 weather factor",MONO,None),
 ("    Navier fare    = comparable premium transfer fare \u00d7 discount (=1.0, market parity)",MONO,None),
 ("EBITDA  = REVENUE \u2212 OPEX",MONO,f_light),
 ("    OPEX = energy + crew + berth/port admin + maintenance + vessel insurance + fast-charge berth",MONO,None),
 ("PAYBACK = vessel CAPEX \u00f7 EBITDA",MONO,f_light),
 ("",None,None),
 ("OPEX lines (per boat, per year \u2014 no double-counting)",BOLD,None),
 ("Energy \u2014 variable propulsion cost: kWh on revenue sailings \u00d7 local $/kWh (Country opex tab). Excludes utility demand charges.",None,None),
 ("Crew \u2014 fully-loaded captain + relief: local captain wage \u00d7 crew FTE factor (Country opex + Assumptions).",None,None),
 ("Berth & port admin \u2014 fixed annual berth, port fees, and local marine admin (Country opex tab). Excludes vessel H&M+P&I and fast-charge infrastructure.",None,None),
 ("Maintenance \u2014 vessel annual maintenance reserve (Assumptions tab).",None,None),
 ("Vessel insurance \u2014 commercial H&M + P&I as % of regional CAPEX (Assumptions % \u00d7 Country opex CAPEX column).",None,None),
 ("Fast-charge berth \u2014 dedicated fast-charge berth slot + utility demand charges (Assumptions default; overridable per market). Separate from per-kWh energy and berth/port admin.",None,None),
 ("",None,None),
 ("From demand pool to fleet & market size",BOLD,None),
 ("Navier serviceable rides/yr = corridor demand pool \u00d7 capture rate (10% on contested corridors; ~90% on captive sole-operator transfer legs \u2014 see the per-row Capture % column)",MONO,None),
 ("Vessels supported = FLOOR(Navier rides/yr \u00f7 pax/yr per boat);  Market rev = vessels \u00d7 rev/boat. The 'Market sizing' tab rolls these up into the SOM \u2192 SAM \u2192 TAM ladder.",None,None),
 ("",None,None),
 ("The scenario toggle (what we flex)",BOLD,None),
 ("'Corridor economics' has a SCENARIO cell = THIN / MID / FULL. It flexes seat occupancy and revenue-leg utilization as separate inputs. South Korea MID uses 65% for each, with gross legs/day schedule-derived and uncapped; thin/full occupancy bookends remain 45%/70%. Other countries retain their existing caps and bands. Payback THIN/MID/FULL columns always show all three at once.",None,None),
 ("",None,None),
 ("Colour & provenance legend",BOLD,None),
 ("   Yellow = INPUTS you can change (fare, distance, demand, weather).",None,f_input),
 ("   Green  = COMPUTED by formula.",None,f_calc),
 ("   Every fare & demand figure carries a Source tier (T1 official \u2192 T5 modeled) + Confidence (high/med/low). Null beats a wrong number: no published pool \u2192 blank, not a guess.",None,None),
 (f"   Economics holds: {len(economics_holds)} corridor(s) are explicitly excluded; see the Economics holds tab. Missing country costs never borrow Singapore or another market.",BOLD,f_band),
]
if ws0 is not None:
    rr=4
    for text,font,fill in content:
        sc(ws0,f"B{rr}",text,font=font,fill=fill,align=wrap); rr+=1

# ------------------------------------------------------------ TAB 1 Assumptions
ws1=wb.create_sheet("Assumptions"); ws1.sheet_view.showGridLines=False
for i,wd in enumerate([30,14,12,8,9,72],1): ws1.column_dimensions[get_column_letter(i)].width=wd
sc(ws1,"A1","Assumptions \u2014 single source of truth (the engine references these named cells)",font=H2,fill=f_navy); ws1.merge_cells("A1:F1")
HDRS=["Parameter","Value","Unit","Tier","Conf.","Source / note"]
def sec(r,title): sc(ws1,f"A{r}",title,font=H2,fill=f_steel); ws1.merge_cells(f"A{r}:F{r}")
def hrow(r):
    for i,h in enumerate(HDRS): sc(ws1,f"{get_column_letter(i+1)}{r}",h,font=HDR,fill=f_steel,align=ctr,bd=True)
def param(r,label,value,unit,tr,cf,source,fmt=None,name=None):
    sc(ws1,f"A{r}",label,bd=True)
    sc(ws1,f"B{r}",value,fill=f_input,fmt=fmt,align=ctr,bd=True,font=BOLD)
    sc(ws1,f"C{r}",unit,align=ctr,bd=True,font=SMALL); sc(ws1,f"D{r}",tr,align=ctr,bd=True,font=SMALL)
    sc(ws1,f"E{r}",cf,align=ctr,bd=True,font=SMALL); sc(ws1,f"F{r}",source,align=wrap,bd=True,font=SMALL)
    if name: addname(name,f"'Assumptions'!$B${r}")
    return r+1

r=3
sec(r,"Vessel \u2014 N30 Pioneer II  (L1 global, commercial now)"); r+=1; hrow(r); r+=1
r=param(r,"Passenger capacity",pax_cap,"pax",tier(P2["pax_capacity"]),conf(P2["pax_capacity"]),src(P2["pax_capacity"]),NUM,"pax_cap")
r=param(r,"Range",range_nm,"nm",tier(P2["range_nm"]),conf(P2["range_nm"]),src(P2["range_nm"]),NUM,"range_nm")
r=param(r,"Cruise speed",cruise_kt,"kt",tier(P2["cruise_speed_kt"]),conf(P2["cruise_speed_kt"]),src(P2["cruise_speed_kt"]),NUM,"cruise_kt")
_capex_label = "CAPEX (per vessel; hospitality N30 list = $1M, region-independent \u2014 see Country opex tab col G)" if HOSPITALITY_CAPEX else "CAPEX (per vessel; US/EU base \u2014 ROW=$600K, see Country opex tab col G)"
r=param(r,_capex_label,capex,"USD",tier(P2["capex_usd"]),conf(P2["capex_usd"]),src(P2["capex_usd"]),USD,"capex")
r=param(r,"Battery",battery,"kWh",tier(P2["battery_kwh"]),conf(P2["battery_kwh"]),src(P2["battery_kwh"]),NUM,"battery")
r=param(r,"Annual maintenance",maint,"USD/yr",tier(P2["annual_maintenance_usd"]),conf(P2["annual_maintenance_usd"]),src(P2["annual_maintenance_usd"]),USD,"maint")
r=param(r,"Depreciation life",dep_years,"yr",tier(P2["depreciation_years"]),conf(P2["depreciation_years"]),src(P2["depreciation_years"]),NUM,"dep_years")
r=param(r,"Diesel comparable economy",diesel_nmpg,"nm/gal",tier(P2["diesel_comparable_nm_per_gal"]),conf(P2["diesel_comparable_nm_per_gal"]),src(P2["diesel_comparable_nm_per_gal"]),NUM2,"diesel_nmpg")
r+=1
sec(r,"Operating defaults  (L1 global; overridable per market)"); r+=1; hrow(r); r+=1
r=param(r,"Service window",service_hr,"hr/day",tier(OPS["service_window_hr_per_day"]),conf(OPS["service_window_hr_per_day"]),src(OPS["service_window_hr_per_day"]),NUM,"service_hr")
r=param(r,"Turnaround / charge per leg",turn_min,"min",tier(OPS["turnaround_charge_min"]),conf(OPS["turnaround_charge_min"]),src(OPS["turnaround_charge_min"]),NUM,"turn_min")
r=param(r,"Boarding / scheduling dwell per leg",dwell_min,"min",tier(OPS["boarding_dwell_min"]),conf(OPS["boarding_dwell_min"]),src(OPS["boarding_dwell_min"]),NUM,"dwell_min")
r=param(r,"Max revenue sailings/day (cap)",max_tpd,"legs/day",tier(OPS["max_trips_per_day"]),conf(OPS["max_trips_per_day"]),src(OPS["max_trips_per_day"]),NUM,"max_tpd")
r=param(r,"Crew FTE factor (x captain wage)",crew_factor,"x",tier(OPS["crew_fte_factor"]),conf(OPS["crew_fte_factor"]),src(OPS["crew_fte_factor"]),NUM2,"crew_factor")
r=param(r,"Mechanical uptime",mech_uptime,"frac",tier(OPS["monthly_operational_capacity"]),conf(OPS["monthly_operational_capacity"]),src(OPS["monthly_operational_capacity"]),PCT,"mech_uptime")
r=param(r,"Navier capture rate",capture,"frac",tier(OPS["navier_capture_rate"]),conf(OPS["navier_capture_rate"]),"Navier wins ~10% of a CONTESTED corridor demand pool (locked). Captive sole-operator transfer legs (captive:true / luxury_charter / hospitality archetypes) carry the captive rate instead \u2014 see per-row Capture %.",PCT,"capture")
cap_cc=OPS.get("captive_capture_rate")
if OPS.get("capture_override_enabled") and cap_cc is not None:
    r=param(r,"Captive capture rate (A\u2032 corridors)",v(cap_cc),"frac","",conf(cap_cc),"Sole-operator water-access-only transfer legs (captive:true or luxury_charter/hospitality archetype): Navier IS the transfer fleet \u2014 capture ~90%. Applied per-row in the Capture % column.",PCT,"capture_captive")
r=param(r,"Discount factor (vs comparable fare)",discount,"frac","T1","high","Market parity \u2014 better product, not cheaper (locked).",NUM2,"discount")
r=param(r,"Diesel CO\u2082 per gallon",diesel_co2pg,"kg/gal",tier(CAR["diesel_kg_co2_per_gal"]),conf(CAR["diesel_kg_co2_per_gal"]),src(CAR["diesel_kg_co2_per_gal"]),NUM2,"diesel_co2pg")
r+=1
sec(r,"OPEX defaults  (global vessel assumptions \u2014 country costs use exact-key rows only)"); r+=1; hrow(r); r+=1
r=param(r,"Vessel insurance \u2014 H&M + P&I (% of CAPEX/yr)",ins_pct,"frac",tier(OPS["insurance_pct_of_capex"]),conf(OPS["insurance_pct_of_capex"]),
      src(OPS["insurance_pct_of_capex"])+". Multiplied by regional CAPEX (Country opex col G). Not included in berth/port admin.",PCT,"ins_pct")
r=param(r,"Fast-charge berth & demand charges (global default)",charge_berth,"USD/yr",tier(OPS["charging_berth_annual_usd"]),conf(OPS["charging_berth_annual_usd"]),
      src(OPS["charging_berth_annual_usd"])+". Dedicated charge berth + utility demand charges. Excludes per-kWh propulsion energy and berth/port admin. L3-overridable per market.",USD,"charge_berth")
r+=1
sec(r,"South Korea schedule / capacity override  (country == 'South Korea' only)"); r+=1; hrow(r); r+=1
r=param(r,"South Korea MID seat occupancy",korea_mid_load,"share of seats",tier(KOREA_OPS["mid_load_factor"]),conf(KOREA_OPS["mid_load_factor"]),src(KOREA_OPS["mid_load_factor"]),PCT,"korea_mid_load")
r=param(r,"South Korea MID revenue-leg utilization",korea_mid_revleg,"share of gross legs",tier(KOREA_OPS["mid_revenue_leg_factor"]),conf(KOREA_OPS["mid_revenue_leg_factor"]),src(KOREA_OPS["mid_revenue_leg_factor"]),PCT,"korea_mid_revleg")
r=param(r,"South Korea service window",korea_service_hr,"hr/day",tier(KOREA_OPS["service_window_hr_per_day"]),conf(KOREA_OPS["service_window_hr_per_day"]),src(KOREA_OPS["service_window_hr_per_day"]),NUM,"korea_service_hr")
r=param(r,"South Korea turnaround",korea_turn_min,"min/gross leg",tier(KOREA_OPS["turnaround_min"]),conf(KOREA_OPS["turnaround_min"]),src(KOREA_OPS["turnaround_min"]),NUM,"korea_turn_min")
r=param(r,"South Korea boarding dwell",korea_dwell_min,"min/gross leg",tier(KOREA_OPS["boarding_dwell_min"]),conf(KOREA_OPS["boarding_dwell_min"]),src(KOREA_OPS["boarding_dwell_min"]),NUM,"korea_dwell_min")
r=param(r,"South Korea charge-planning range basis",korea_charge_range_nm,"nm",korea_charge_cfg.get("source_tier",""),korea_charge_cfg.get("confidence",""),korea_charge_cfg.get("source","")+" "+korea_charge_cfg.get("engineering_status",""),NUM,"korea_charge_range_nm")
r=param(r,"South Korea full-range charge-planning time",korea_full_range_charge_min,"min",korea_charge_cfg.get("source_tier",""),korea_charge_cfg.get("confidence",""),"PLANNING PROXY: charge recovery = distance_nm / 70 × 45. Requires engineering validation; not a certified charge curve.",NUM,"korea_full_range_charge_min")
r+=1
# scenario bands
sec(r,"Utilization scenarios  (seat occupancy + revenue-leg utilization)"); r+=1
scen_hdr_r=r
for i,h in enumerate(["Scenario","Seat occupancy","Revenue-leg utilization","","","Note"]): sc(ws1,f"{get_column_letter(i+1)}{r}",h,font=HDR,fill=f_steel,align=ctr,bd=True)
r+=1
band_rows={}
for b in ("thin","mid","full"):
    lf,rl=bands[b]
    sc(ws1,f"A{r}",b.upper(),bd=True,font=BOLD,align=ctr)
    sc(ws1,f"B{r}",lf,fill=f_input,fmt=NUM2,align=ctr,bd=True); addname(f"load_{b}",f"'Assumptions'!$B${r}")
    sc(ws1,f"C{r}",rl,fill=f_input,fmt=NUM2,align=ctr,bd=True); addname(f"revleg_{b}",f"'Assumptions'!$C${r}")
    sc(ws1,f"F{r}",SCEN[b]["note"],align=wrap,bd=True,font=SMALL)
    band_rows[b]=r; r+=1
band_tbl=f"'Assumptions'!$A${band_rows['thin']}:$C${band_rows['full']}"
addname("band_tbl", band_tbl)
r+=1
# scenario selector + resolved levers
sc(ws1,f"A{r}","SCENARIO (toggle me)",font=BOLD,fill=f_band,bd=True)
sel=sc(ws1,f"B{r}","MID",fill=f_band,align=ctr,bd=True,font=Font(bold=True,size=12)); addname("scenario",f"'Assumptions'!$B${r}")
dv=DataValidation(type="list",formula1='"THIN,MID,FULL"',allow_blank=False); ws1.add_data_validation(dv); dv.add(sel)
sc(ws1,f"C{r}","\u2190 set THIN / MID / FULL",font=SMALL,align=Alignment(vertical="center")); r+=1
sc(ws1,f"A{r}","Selected global seat occupancy",bd=True)
sc(ws1,f"B{r}",f'=VLOOKUP(scenario,{band_tbl},2,FALSE)',fill=f_calc,fmt=NUM2,align=ctr,bd=True); addname("load_sel",f"'Assumptions'!$B${r}"); r+=1
sc(ws1,f"A{r}","Selected global revenue-leg utilization",bd=True)
sc(ws1,f"B{r}",f'=VLOOKUP(scenario,{band_tbl},3,FALSE)',fill=f_calc,fmt=NUM2,align=ctr,bd=True); addname("revleg_sel",f"'Assumptions'!$B${r}"); r+=1

# ------------------------------------------------------------ TAB 2 Country opex
# LB-243 (Jaideep 2026-06-19): per-country CAPEX rule rendered as a sheet column.
#   US + EU member states -> $900K; every other market -> $600K. Mirrors
#   model/aggregate.py capex_for_country() so the transparent sheet recomputes payback honestly.
_EU_COUNTRIES = {
    "Austria","Belgium","Bulgaria","Croatia","Cyprus","Czechia","Czech Republic","Denmark",
    "Estonia","Finland","France","Germany","Greece","Hungary","Ireland","Italy","Latvia",
    "Lithuania","Luxembourg","Malta","Netherlands","Poland","Portugal","Romania","Slovakia",
    "Slovenia","Spain","Sweden",
}
_US_COUNTRIES = {"United States","USA","United States of America"}
def _capex_for_country(ct):
    if HOSPITALITY_CAPEX:                       # LB-260: hospitality N30 list price, region-independent
        return 1000000
    return 900000 if (ct in _EU_COUNTRIES or ct in _US_COUNTRIES) else 600000
ws2=wb.create_sheet("Country opex"); ws2.sheet_view.showGridLines=False
for i,wd in enumerate([18,16,14,14,14,11,14,60],1): ws2.column_dimensions[get_column_letter(i)].width=wd
sc(ws2,"A1","Country opex \u2014 localized inputs (crew, energy, berth admin, CAPEX)",font=H2,fill=f_navy); ws2.merge_cells("A1:H1")
ch=["Country","Captain $/yr","Energy $/kWh","Grid CO\u2082 kg/kWh","Berth & port admin $/yr","Cost index","CAPEX $/vessel","Source notes (crew | energy | berth admin)"]
hr2=3
for i,h in enumerate(ch): sc(ws2,f"{get_column_letter(i+1)}{hr2}",h,font=HDR,fill=f_steel,align=ctrw,bd=True)
cr=hr2+1; country_first=cr
for ct in used_countries:
    row=cref[ct]
    sc(ws2,f"A{cr}",ct,bd=True,font=BOLD)
    sc(ws2,f"B{cr}",v(row["captain_usd_yr"]),fill=f_input,fmt=USD,align=ctr,bd=True)
    sc(ws2,f"C{cr}",v(row["energy_usd_kwh"]),fill=f_input,fmt=USD2,align=ctr,bd=True)
    sc(ws2,f"D{cr}",v(row["grid_co2_kg_kwh"]),fill=f_input,fmt=NUM2,align=ctr,bd=True)
    sc(ws2,f"E{cr}",v(row["marina_overhead_usd_yr"]),fill=f_input,fmt=USD,align=ctr,bd=True)
    sc(ws2,f"F{cr}",v(row.get("cost_index",{})),fill=f_input,fmt=NUM2,align=ctr,bd=True)
    sc(ws2,f"G{cr}",_capex_for_country(ct),fill=f_input,fmt=USD,align=ctr,bd=True)
    _capex_note = "CAPEX LB-260 hospitality N30 list = $1M (region-independent)" if HOSPITALITY_CAPEX else "CAPEX LB-243 US/EU=$900K else $600K"
    note=f"{src(row['captain_usd_yr'])} | {src(row['energy_usd_kwh'])} | {src(row['marina_overhead_usd_yr'])} | {_capex_note}"
    sc(ws2,f"H{cr}",note,align=wrap,bd=True,font=SMALL)
    cr+=1
country_last=cr-1
copex=f"'Country opex'!$A${country_first}:$G${country_last}"
addname("country_opex",copex)
# column index map for VLOOKUP: A=1 country,2 captain,3 energy,4 grid,5 berth/port admin,6 costidx,7 capex

# ------------------------------------------------------------ TAB 3 Economics holds (fail-closed ledger)
wsh=wb.create_sheet("Economics holds"); wsh.sheet_view.showGridLines=False
for i,wd in enumerate([24,58,20,24,86],1): wsh.column_dimensions[get_column_letter(i)].width=wd
sc(wsh,"A1","Economics holds — excluded before model and sheet calculation",font=H2,fill=f_navy); wsh.merge_cells("A1:E1")
hh=["Market","Corridor","Route ID","Country label","Hold reason"]
for i,h in enumerate(hh): sc(wsh,f"{get_column_letter(i+1)}3",h,font=HDR,fill=f_steel,align=ctrw,bd=True)
for rr,h in enumerate(economics_holds,4):
    for cc,key in enumerate(("market","corridor","route_id","country","reason"),1):
        sc(wsh,f"{get_column_letter(cc)}{rr}",h.get(key),align=wrap,bd=True,font=SMALL if cc>1 else BOLD)
if not economics_holds:
    sc(wsh,"A4","None — all in-scope corridors have complete exact-key country references.",font=BOLD,fill=f_calc); wsh.merge_cells("A4:E4")

# ------------------------------------------------------------ TAB 4 Corridor economics (engine)
ws=wb.create_sheet("Corridor economics"); ws.sheet_view.showGridLines=False
# scenario toggle mirror at top
_corr_hdr = ("Corridor economics \u2014 unique geometry (one row per pier-pair; canonical list for Global TAM)"
             if (GLOBAL_MODE and DEDUP_MODE == "unique") else
             f"Navier \u00d7 {PARTNER_LABEL} \u2014 Corridor economics (one boat, one year; MID = headline)")
sc(ws,"A1",_corr_hdr,font=H2,fill=f_navy)
ws.merge_cells("A1:L1")
sc(ws,"N1","SCENARIO:",font=BOLD,align=Alignment(horizontal="right",vertical="center"))
sc(ws,"O1","=scenario",fill=f_band,align=ctr,font=Font(bold=True,size=12),bd=True)
sc(ws,"P1","(set on Assumptions tab)",font=SMALL,align=Alignment(vertical="center"))

# column plan
_cols_head = [
 ("Market","market",10,None,"in"),
 ("Country","country",13,None,"in"),
 ("Corridor","corridor",34,None,"in"),
]
if GLOBAL_MODE and DEDUP_MODE == "unique":
    _cols_head += [
        ("Provenance partner","provenance",14,None,"in"),
        ("Alias markets","alias_markets",30,None,"in"),
        ("Alias count","alias_count",8,NUM,"in"),
    ]
cols = _cols_head + [
 ("Distance","nm",8,NUM,"in"),
 ("Premium fare $/pax","fare",11,USD,"in"),
 ("Fare tier","fare_tier",7,None,"in"),
 ("Fare conf.","fare_conf",8,None,"in"),
 ("Weather factor","weather",9,NUM2,"in"),
 ("One-way min","",8,NUM1,"f"),
 ("Charge recovery min","",10,NUM1,"f"),
 ("Cycle min","",8,NUM1,"f"),
 ("Gross legs/day","",10,NUM,"f"),
 ("Op days/yr","",8,NUM,"f"),
 ("Mech uptime","",8,PCT,"f"),
 ("Revenue-leg utilization","",11,PCT,"f"),
 ("Revenue legs/yr","",10,NUM,"f"),
 ("Seat occupancy","",9,PCT,"f"),
 ("Pax/trip","",8,NUM2,"f"),
 ("Pax/yr","",9,NUM,"f"),
 ("Navier fare $","",9,USD2,"f"),
 ("Revenue/yr","",11,USD,"f"),
 ("Energy/yr","",10,USD,"f"),
 ("Crew/yr","",10,USD,"f"),
 ("Berth & port admin/yr","",12,USD,"f"),
 ("Maint/yr","",9,USD,"f"),
 ("Vessel insurance/yr","",12,USD,"f"),
 ("Fast-charge berth/yr","",12,USD,"f"),
 ("OPEX/yr","",11,USD,"f"),
 ("Deprec/yr","",10,USD,"f"),
 ("EBITDA/yr","",11,USD,"f"),
 ("Margin","",8,PCT,"f"),
 ("Payback yr (sel)","",9,NUM2,"f"),
 ("CO\u2082 saved t/yr","",9,NUM1,"f"),
 ("Demand pool 1-way/yr","demand",12,NUM,"in"),
 ("Demand tier","demand_tier",7,None,"in"),
 ("Demand conf.","demand_conf",8,None,"in"),
 ("Capture %","cap",8,PCT,"in"),
 ("Navier rides/yr","",10,NUM,"f"),
 ("Vessels @capture","",9,NUM,"f"),
 ("Market rev/yr","",13,USD,"f"),
 ("Vessels raw (unfloored)","",9,NUM2,"f"),
 ("Market rev/yr raw","",13,USD,"f"),
 ("Subset of (excl. from market fleet)","subset",26,None,"in"),
 ("Floor bucket (aggregate status)","floor_bucket",14,None,"in"),
 ("Forward SAM (2030-dated; held OUT of grounded floor)","fwd",16,None,"in"),
 ("Upside tier (LB-99; held OUT of grounded floor)","tier_excl",16,None,"in"),
 ("Fleet ceiling (captive villas/25)","fleet_cap",9,NUM,"in"),
 ("Payback THIN","",9,NUM2,"b"),
 ("Payback MID","",9,NUM2,"b"),
 ("Payback FULL","",9,NUM2,"b"),
 ("Fare source (provenance)","fare_src",50,None,"in"),
 ("Demand source (provenance)","demand_src",50,None,"in"),
 ("Route ID","route_id",16,None,"in"),
]
# map header -> column letter
colletter={}
for i,(h,_,wd,_,_) in enumerate(cols):
    L=get_column_letter(i+1); colletter[h]=L; ws.column_dimensions[L].width=wd
HROW=3
for i,(h,key,wd,fmt,kind) in enumerate(cols):
    L=get_column_letter(i+1)
    fill=f_steel
    sc(ws,f"{L}{HROW}",h,font=HDR,fill=fill,align=ctrw,bd=True)
ws.row_dimensions[HROW].height=42

def CL(name): return colletter[name]
DATA0=HROW+1
for idx,rec in enumerate(rows):
    R=DATA0+idx
    def put(name,val,fill,fmt=None,font=None,align=None):
        sc(ws,f"{CL(name)}{R}",val,fill=fill,fmt=fmt,font=font or SMALL if fill is f_input else font,align=align,bd=True)
    # inputs
    sc(ws,f"{CL('Market')}{R}",rec["market"],fill=f_input,bd=True,font=SMALL)
    sc(ws,f"{CL('Country')}{R}",rec["country"],fill=f_input,bd=True,font=SMALL)
    sc(ws,f"{CL('Corridor')}{R}",rec["corridor"],fill=f_input,bd=True,font=Font(size=9),align=wrap)
    if GLOBAL_MODE and DEDUP_MODE == "unique":
        sc(ws,f"{CL('Provenance partner')}{R}",rec.get("provenance"),fill=f_input,bd=True,font=SMALL)
        sc(ws,f"{CL('Alias markets')}{R}",rec.get("alias_markets"),fill=f_grey,align=wrap,bd=True,font=Font(size=8,color="555555"))
        sc(ws,f"{CL('Alias count')}{R}",rec.get("alias_count", 1),fill=f_grey,fmt=NUM,align=ctr,bd=True,font=SMALL)
    sc(ws,f"{CL('Distance')}{R}",rec["nm"],fill=f_input,fmt=NUM,align=ctr,bd=True)
    sc(ws,f"{CL('Premium fare $/pax')}{R}",rec["fare"],fill=f_input,fmt=USD,align=ctr,bd=True)
    sc(ws,f"{CL('Fare tier')}{R}",rec["fare_tier"],fill=f_input,align=ctr,bd=True,font=SMALL)
    sc(ws,f"{CL('Fare conf.')}{R}",rec["fare_conf"],fill=f_input,align=ctr,bd=True,font=SMALL)
    sc(ws,f"{CL('Weather factor')}{R}",rec["weather"],fill=f_input,fmt=NUM2,align=ctr,bd=True)
    nm=f"{CL('Distance')}{R}"; fare=f"{CL('Premium fare $/pax')}{R}"; wf=f"{CL('Weather factor')}{R}"; ctry=f"{CL('Country')}{R}"
    ow=f"{CL('One-way min')}{R}"; charge=f"{CL('Charge recovery min')}{R}"; cyc=f"{CL('Cycle min')}{R}"; tpd=f"{CL('Gross legs/day')}{R}"
    upt=f"{CL('Mech uptime')}{R}"; rlp=f"{CL('Revenue-leg utilization')}{R}"
    od=f"{CL('Op days/yr')}{R}"; tpy=f"{CL('Revenue legs/yr')}{R}"; lf=f"{CL('Seat occupancy')}{R}"
    ppt=f"{CL('Pax/trip')}{R}"; ppy=f"{CL('Pax/yr')}{R}"; nf=f"{CL('Navier fare $')}{R}"
    rev=f"{CL('Revenue/yr')}{R}"; en=f"{CL('Energy/yr')}{R}"; cap=f"{CL('Crew/yr')}{R}"
    mar=f"{CL('Berth & port admin/yr')}{R}"; mnt=f"{CL('Maint/yr')}{R}"; opx=f"{CL('OPEX/yr')}{R}"
    insr=f"{CL('Vessel insurance/yr')}{R}"; chg=f"{CL('Fast-charge berth/yr')}{R}"
    dep=f"{CL('Deprec/yr')}{R}"; ebt=f"{CL('EBITDA/yr')}{R}"; mgn=f"{CL('Margin')}{R}"
    _co2_lbl = "CO\u2082 saved t/yr"
    pbk=f"{CL('Payback yr (sel)')}{R}"; co2=f"{CL(_co2_lbl)}{R}"
    dem=f"{CL('Demand pool 1-way/yr')}{R}"; nrd=f"{CL('Navier rides/yr')}{R}"
    capc=f"{CL('Capture %')}{R}"
    ves=f"{CL('Vessels @capture')}{R}"; mrev=f"{CL('Market rev/yr')}{R}"
    # formulas (computed, green)
    def F(ref,formula,fmt=None,align=ctr):
        sc(ws,ref,formula,fill=f_calc,fmt=fmt,align=align,bd=True,font=Font(size=9))
    F(ow,f"=60*{nm}/cruise_kt",NUM1)
    F(charge,f'=IF({ctry}="{KOREA_COUNTRY}",{nm}/korea_charge_range_nm*korea_full_range_charge_min,0)',NUM1)
    F(cyc,f'=IF({ctry}="{KOREA_COUNTRY}",{ow}+korea_turn_min+korea_dwell_min+{charge},{ow}+turn_min+dwell_min)',NUM1)
    F(tpd,f'=IF({ctry}="{KOREA_COUNTRY}",FLOOR(60*korea_service_hr/{cyc},1),MIN({rec.get("eff_cap",max_tpd)},FLOOR(60*service_hr/{cyc},1)))',NUM)
    if rec.get("season_days") is not None:
        sc(ws,od,rec["season_days"],fill=f_input,fmt=NUM,align=ctr,bd=True)
    else:
        # atom.py: season_days = round(365 * op_capacity * weather)  → half-even
        F(od,f"={round_half_even(f'365*mech_uptime*{wf}')}",NUM)
    F(upt,"=mech_uptime",PCT)
    F(rlp,f'=IF(AND({ctry}="{KOREA_COUNTRY}",scenario="MID"),korea_mid_revleg,revleg_sel)',PCT)
    # atom.py: revenue legs/year = round(gross legs/day * season days * revenue-leg utilization)
    F(tpy,f"={round_half_even(f'{tpd}*{od}*{rlp}')}",NUM)
    F(lf,f'=IF(AND({ctry}="{KOREA_COUNTRY}",scenario="MID"),korea_mid_load,load_sel)',PCT)
    F(ppt,f"=pax_cap*{lf}",NUM2)
    F(ppy,f"={ppt}*{tpy}",NUM)
    F(nf,f"={fare}*discount",USD2)
    F(rev,f"={nf}*{ppy}",USD)
    # opex
    F(en,f"=(battery/range_nm)*{nm}*{tpy}*VLOOKUP({ctry},country_opex,3,FALSE)",USD)
    F(cap,f"=VLOOKUP({ctry},country_opex,2,FALSE)*crew_factor",USD)
    F(mar,f"=VLOOKUP({ctry},country_opex,5,FALSE)",USD)
    F(mnt,"=maint",USD)
    cpx=f"VLOOKUP({ctry},country_opex,7,FALSE)"  # LB-243 per-country capex
    F(insr,f"={cpx}*ins_pct",USD)
    F(chg,"=charge_berth",USD)
    F(opx,f"={en}+{cap}+{mar}+{mnt}+{insr}+{chg}",USD)
    F(dep,f"={cpx}/dep_years",USD)
    F(ebt,f"={rev}-{opx}",USD)
    F(mgn,f"=IF({rev}=0,\"\",{ebt}/{rev})",PCT)
    F(pbk,f"=IF({ebt}<=0,\"\",{cpx}/{ebt})",NUM2)
    F(co2,f"=(({nm}*{tpy}/diesel_nmpg)*diesel_co2pg-(battery/range_nm)*{nm}*{tpy}*VLOOKUP({ctry},country_opex,4,FALSE))/1000",NUM1)
    # demand inputs
    sc(ws,dem,rec["demand"],fill=f_input,fmt=NUM,align=ctr,bd=True)
    sc(ws,f"{CL('Demand tier')}{R}",rec["demand_tier"],fill=f_input,align=ctr,bd=True,font=SMALL)
    sc(ws,f"{CL('Demand conf.')}{R}",rec["demand_conf"],fill=f_input,align=ctr,bd=True,font=SMALL)
    sc(ws,capc,rec["cap"],fill=f_input,fmt=PCT,align=ctr,bd=True)
    F(nrd,f"=IF({dem}=\"\",\"\",{dem}*{capc})",NUM)
    # LB-88: captive-resort fleet ceiling caps vessels (floored AND raw); rev scales to cap
    fcap=f"{CL('Fleet ceiling (captive villas/25)')}{R}"
    F(ves,f"=IF(OR({dem}=\"\",{ppy}=0),\"\",MIN(IF({fcap}=\"\",9.9E+99,{fcap}),FLOOR({nrd}/{ppy},1)))",NUM)
    # atom.py: rev_per_year = round(navier_fare * pax_per_year, 0); market_rev uses
    # that whole-dollar per-boat revenue × floored vessels. Half-even for .5 ties.
    F(mrev,f"=IF({ves}=\"\",\"\",{ves}*{round_half_even(rev)})",USD)
    # R-FLOOR-2 / G51 network-sum basis. South Korea mirrors atom.py exactly:
    # unfloored vessel fraction × whole-dollar (half-even) per-boat revenue.
    # Other countries retain the prior unrounded sheet behavior (no blast radius).
    vraw=f"{CL('Vessels raw (unfloored)')}{R}"; mrevraw=f"{CL('Market rev/yr raw')}{R}"
    F(vraw,f"=IF(OR({dem}=\"\",{ppy}=0),\"\",MIN(IF({fcap}=\"\",9.9E+99,{fcap}),{nrd}/{ppy}))",NUM2)
    F(mrevraw,f'=IF({vraw}="","",{vraw}*IF({ctry}="{KOREA_COUNTRY}",{round_half_even(rev)},{rev}))',USD)
    sc(ws,f"{CL('Subset of (excl. from market fleet)')}{R}",rec["subset"],fill=f_input,align=wrap,bd=True,font=Font(size=8,color="555555"))
    # LB-33 forward-SAM flag: 2030-dated / low-confidence demand — engine row computed
    # at MID but held OUT of the grounded floor (separate FORWARD SAM total below).
    sc(ws,f"{CL('Floor bucket (aggregate status)')}{R}",rec["floor_bucket"],fill=f_input,align=ctr,bd=True,font=SMALL)
    sc(ws,f"{CL('Forward SAM (2030-dated; held OUT of grounded floor)')}{R}",rec["fwd"],fill=f_input,align=ctr,bd=True,font=Font(size=8,color="9C4500"))
    sc(ws,f"{CL('Upside tier (LB-99; held OUT of grounded floor)')}{R}",rec["tier_excl"],fill=f_input,align=ctr,bd=True,font=Font(size=8,color="9C4500"))
    sc(ws,f"{CL('Fleet ceiling (captive villas/25)')}{R}",rec["fleet_cap"],fill=f_input,fmt=NUM,align=ctr,bd=True,font=SMALL)
    # band paybacks (self-contained per band) — trips/yr half-even like selected scenario
    for b,name in (("thin","Payback THIN"),("mid","Payback MID"),("full","Payback FULL")):
        revleg_b=(f'IF({ctry}="{KOREA_COUNTRY}",korea_mid_revleg,revleg_mid)' if b == "mid" else f"revleg_{b}")
        load_b=(f'IF({ctry}="{KOREA_COUNTRY}",korea_mid_load,load_mid)' if b == "mid" else f"load_{b}")
        tpyb=round_half_even(f"{tpd}*{od}*{revleg_b}")
        revb=f"({nf}*pax_cap*{load_b}*{tpyb})"
        enb=f"((battery/range_nm)*{nm}*{tpyb}*VLOOKUP({ctry},country_opex,3,FALSE))"
        opxb=f"({enb}+{cap}+{mar}+{mnt}+{insr}+{chg})"
        ebtb=f"({revb}-{opxb})"
        sc(ws,f"{CL(name)}{R}",f"=IF({ebtb}<=0,\"\",{cpx}/{ebtb})",fill=f_band,fmt=NUM2,align=ctr,bd=True,font=Font(size=9))
    # provenance
    sc(ws,f"{CL('Fare source (provenance)')}{R}",rec["fare_src"],fill=f_grey,align=wrap,bd=True,font=Font(size=8,color="555555"))
    sc(ws,f"{CL('Demand source (provenance)')}{R}",rec["demand_src"],fill=f_grey,align=wrap,bd=True,font=Font(size=8,color="555555"))
    sc(ws,f"{CL('Route ID')}{R}",rec.get("route_id"),fill=f_input,align=ctr,bd=True,font=SMALL)

LASTROW=DATA0+len(rows)-1
# totals row
TR=LASTROW+1
sc(ws,f"{CL('Market')}{TR}","TOTAL / median",font=BOLD,fill=f_light,bd=True)
sc(ws,f"{CL('Revenue/yr')}{TR}",f"=SUM({CL('Revenue/yr')}{DATA0}:{CL('Revenue/yr')}{LASTROW})",font=BOLD,fill=f_light,fmt=USD,align=ctr,bd=True)
sc(ws,f"{CL('EBITDA/yr')}{TR}",f"=SUM({CL('EBITDA/yr')}{DATA0}:{CL('EBITDA/yr')}{LASTROW})",font=BOLD,fill=f_light,fmt=USD,align=ctr,bd=True)
sc(ws,f"{CL('Margin')}{TR}",f"=IF({CL('Revenue/yr')}{TR}=0,\"\",{CL('EBITDA/yr')}{TR}/{CL('Revenue/yr')}{TR})",font=BOLD,fill=f_light,fmt=PCT,align=ctr,bd=True)
sc(ws,f"{CL('Vessels @capture')}{TR}",f"=SUM({CL('Vessels @capture')}{DATA0}:{CL('Vessels @capture')}{LASTROW})",font=BOLD,fill=f_light,fmt=NUM,align=ctr,bd=True)
sc(ws,f"{CL('Market rev/yr')}{TR}",f"=SUM({CL('Market rev/yr')}{DATA0}:{CL('Market rev/yr')}{LASTROW})",font=BOLD,fill=f_light,fmt=USD,align=ctr,bd=True)
ws.freeze_panes="D4"

# ------------------------------------------------------------ Market fleet build-up
# R-FLOOR-2 / G51 locked ruling: markets flagged fleet_basis="network_sum" size their
# fleet as the SUM of unfloored per-corridor vessel fractions, rounded ONCE per market
# (ceil for grab); corridors marked _subset_of (mint-slices of a parent crossing) are
# EXCLUDED from the market sum. per_corridor_floor markets keep the legacy floored sum.
BU0=TR+2
sc(ws,f"A{BU0}","Market fleet build-up \u2014 grounded floor (network-sum of raw vessel fractions per market, rounded once; subset slices excluded; legacy per-corridor-floor totals above kept for reference)",font=H2,fill=f_navy)
ws.merge_cells(f"A{BU0}:L{BU0}")
buh=BU0+1
for i,h in enumerate(["Market","Fleet basis","Raw vessels (sum)","Fleet (rounded once)","Market rev/yr (raw sum)"]):
    sc(ws,f"{get_column_letter(i+1)}{buh}",h,font=HDR,fill=f_steel,align=ctrw,bd=True)
MKTcol=CL('Market'); RAWcol=CL('Vessels raw (unfloored)'); RRVcol=CL('Market rev/yr raw')
VEScol=CL('Vessels @capture'); MRVcol=CL('Market rev/yr'); SUBcol=CL('Subset of (excl. from market fleet)')
FLOORcol=CL('Floor bucket (aggregate status)')
FWDcol=CL('Forward SAM (2030-dated; held OUT of grounded floor)')
TIERcol=CL('Upside tier (LB-99; held OUT of grounded floor)')
_GROUNDED_CRIT=f',{FLOORcol}{DATA0}:{FLOORcol}{LASTROW},"Grounded"'
bur=buh+1
seen_mkts=[]
for rec in rows:
    if rec["market"] not in seen_mkts: seen_mkts.append(rec["market"])
for mname in seen_mkts:
    basis,rounding=mkt_fleet_basis.get(mname,("per_corridor_floor",None))
    crit=f'"{mname}"'
    sc(ws,f"A{bur}",mname,bd=True,font=BOLD)
    sc(ws,f"B{bur}",f"{basis}{' / '+rounding if rounding else ''}",bd=True,font=SMALL,align=ctr)
    # LB-258: only status=grounded corridors build the published floor (matches growth.py / deck).
    if basis=="network_sum":
        sc(ws,f"C{bur}",f'=SUMIFS({RAWcol}{DATA0}:{RAWcol}{LASTROW},{MKTcol}{DATA0}:{MKTcol}{LASTROW},{crit},{SUBcol}{DATA0}:{SUBcol}{LASTROW},"="{_GROUNDED_CRIT})',fill=f_calc,fmt=NUM2,align=ctr,bd=True)
        rndfn="CEILING" if rounding=="ceil" else "ROUND"
        sc(ws,f"D{bur}",f"=IF(C{bur}=0,0,{rndfn}(C{bur},1))" if rounding=="ceil" else f"=ROUND(C{bur},0)",fill=f_calc,fmt=NUM,align=ctr,bd=True)
        sc(ws,f"E{bur}",f'=SUMIFS({RRVcol}{DATA0}:{RRVcol}{LASTROW},{MKTcol}{DATA0}:{MKTcol}{LASTROW},{crit},{SUBcol}{DATA0}:{SUBcol}{LASTROW},"="{_GROUNDED_CRIT})',fill=f_calc,fmt=USD,align=ctr,bd=True)
    else:
        sc(ws,f"C{bur}","\u2014",align=ctr,bd=True,font=SMALL)
        sc(ws,f"D{bur}",f'=SUMIFS({VEScol}{DATA0}:{VEScol}{LASTROW},{MKTcol}{DATA0}:{MKTcol}{LASTROW},{crit},{SUBcol}{DATA0}:{SUBcol}{LASTROW},"="{_GROUNDED_CRIT})',fill=f_calc,fmt=NUM,align=ctr,bd=True)
        sc(ws,f"E{bur}",f'=SUMIFS({MRVcol}{DATA0}:{MRVcol}{LASTROW},{MKTcol}{DATA0}:{MKTcol}{LASTROW},{crit},{SUBcol}{DATA0}:{SUBcol}{LASTROW},"="{_GROUNDED_CRIT})',fill=f_calc,fmt=USD,align=ctr,bd=True)
    bur+=1
BUT=bur
sc(ws,f"A{BUT}","GROUNDED FLOOR (PUBLISHED)",font=BOLD,fill=f_light,bd=True)
sc(ws,f"D{BUT}",f"=SUM(D{buh+1}:D{BUT-1})",font=BOLD,fill=f_light,fmt=NUM,align=ctr,bd=True)
sc(ws,f"E{BUT}",f"=SUM(E{buh+1}:E{BUT-1})",font=BOLD,fill=f_light,fmt=USD,align=ctr,bd=True)
addname("som_floor_fleet", f"'Corridor economics'!$D${BUT}")
addname("som_floor_rev", f"'Corridor economics'!$E${BUT}")
# ---- FORWARD SAM bucket (LB-33): 2030-dated / low-confidence corridors, per-corridor
# floored fleet (mirrors aggregate.py forward_sam rollup), reported BELOW the floor,
# never summed into it. For all-forward partners (Saudi-PIF / Red Sea Global) the
# Market-sizing ladder anchors HERE — clearly labelled as forward, not grounded.
# LB-99 estimated-upside bucket: modal-shift / experience-upside tiers, reported
# BELOW the floor, never summed into it (mirrors aggregate.py estimated_upside).
ESTROW=None
if any(rec.get("floor_bucket") == "Estimated" for rec in rows):
    ESTROW=BUT+1
    sc(ws,f"A{ESTROW}","ESTIMATED DEMAND (country-median cascade; held OUT of grounded floor)",font=BOLD,fill=f_band,bd=True)
    sc(ws,f"D{ESTROW}",f'=SUMIF({FLOORcol}{DATA0}:{FLOORcol}{LASTROW},"Estimated",{VEScol}{DATA0}:{VEScol}{LASTROW})',font=BOLD,fill=f_band,fmt=NUM,align=ctr,bd=True)
    sc(ws,f"E{ESTROW}",f'=SUMIF({FLOORcol}{DATA0}:{FLOORcol}{LASTROW},"Estimated",{MRVcol}{DATA0}:{MRVcol}{LASTROW})',font=BOLD,fill=f_band,fmt=USD,align=ctr,bd=True)
UPROW=None
if any(rec.get("floor_bucket") == "Upside tier" for rec in rows):
    UPROW=(ESTROW or BUT)+1
    sc(ws,f"A{UPROW}","ESTIMATED UPSIDE (LB-99 modal-shift / experience-upside tiers; held OUT of grounded floor)",font=BOLD,fill=f_band,bd=True)
    sc(ws,f"D{UPROW}",f'=SUMIF({FLOORcol}{DATA0}:{FLOORcol}{LASTROW},"Upside tier",{VEScol}{DATA0}:{VEScol}{LASTROW})',font=BOLD,fill=f_band,fmt=NUM,align=ctr,bd=True)
    sc(ws,f"E{UPROW}",f'=SUMIF({FLOORcol}{DATA0}:{FLOORcol}{LASTROW},"Upside tier",{MRVcol}{DATA0}:{MRVcol}{LASTROW})',font=BOLD,fill=f_band,fmt=USD,align=ctr,bd=True)
FSROW=None
if any(rec.get("floor_bucket") == "Forward SAM" for rec in rows):
    FSROW=(UPROW or ESTROW or BUT)+1
    sc(ws,f"A{FSROW}","FORWARD SAM (2030-dated; held OUT of grounded floor)",font=BOLD,fill=f_band,bd=True)
    sc(ws,f"D{FSROW}",f'=SUMIF({FLOORcol}{DATA0}:{FLOORcol}{LASTROW},"Forward SAM",{VEScol}{DATA0}:{VEScol}{LASTROW})',font=BOLD,fill=f_band,fmt=NUM,align=ctr,bd=True)
    sc(ws,f"E{FSROW}",f'=SUMIF({FLOORcol}{DATA0}:{FLOORcol}{LASTROW},"Forward SAM",{MRVcol}{DATA0}:{MRVcol}{LASTROW})',font=BOLD,fill=f_band,fmt=USD,align=ctr,bd=True)
ALL_FWD = bool(rows) and all(rec.get("fwd") for rec in rows)
if ALL_FWD:
    # re-anchor the ladder named ranges to the forward-SAM bucket (grounded floor is 0 by design)
    named["som_floor_fleet"]=f"'Corridor economics'!$D${FSROW}"
    named["som_floor_rev"]=f"'Corridor economics'!$E${FSROW}"

# ------------------------------------------------------------ TAB 4 Market sizing / Global TAM
_ms_title=("Market sizing \u2014 how the FORWARD-SAM anchor (2030-dated; no near-term grounded floor) builds into SOM \u2192 SAM \u2192 TAM"
           if ALL_FWD else
           "Market sizing \u2014 how the grounded floor builds into SOM \u2192 SAM \u2192 TAM")
if GLOBAL_MODE and DEDUP_MODE == "unique":
    _ms_title = "Global TAM \u2014 unique geometry (one row per pier-pair; SOM \u2192 SAM \u2192 TAM)"
ws4_title = "Global TAM" if (GLOBAL_MODE and DEDUP_MODE == "unique") else "Market sizing"
ws4=wb.create_sheet(ws4_title); ws4.sheet_view.showGridLines=False
for i,wd in enumerate([40,16,16,16,72],1): ws4.column_dimensions[get_column_letter(i)].width=wd
sc(ws4,"A1",_ms_title,font=H2,fill=f_navy); ws4.merge_cells("A1:E1")
M=gcfg["multipliers"]; LAD=gcfg["ladder"]
# LB-172: Partner-scoped multiplier overrides. Partners whose corridor census already
# covers the entire geographically addressable network (small island states, atoll-bounded
# systems) get greenfield=1.0 — no greenfield extension possible. Drops Marine TAM from
# global $3.5B template to the deck-locked partner cap.
PARTNER_MULTIPLIER_OVERRIDES = {
    "french-polynesia": {
        "greenfield_corridor_factor": {"low":1.0, "mid":1.0, "high":1.0,
            "_doc":"FP geographically bounded — 23 sourced corridors ≈ full addressable network. No greenfield extension."},
    },
    # LB-196 / econ-4b: Careem uses its OWN UAE census (#79ae corpus), NOT Grab's default
    # growth-config greenfield (3.44/4.9/6.36). Rebuilt on the post-P1-scrub recalibrated
    # corpus: 18 tier-1 greenfield vs 35 sourced → 1 + (18/35)*alpha. Reconciles the sheet
    # ladder to growth.py --partner careem --greenfield-json careem-greenfield-census.json.
    "careem": {
        "greenfield_corridor_factor": {"low":1.129, "mid":1.206, "high":1.283,
            "_doc":"UAE census (#79ae): 18 tier-1 market-making greenfield corridors vs 35 sourced, alpha 0.25/0.40/0.55. Recal scrubbed the pre-P1 noise corridors → honestly thinner than Grab's 9.7× ratio."},
    },
    # LB-249 / LB-250 (Jaideep 2026-06-19): Bolt & Yango have no partner-specific greenfield
    # census yet, but our sourced network demonstrably under-counts the real addressable web of
    # crossings, so greenfield stays ON. Per Jaideep's call we apply the GLOBAL TEMPLATE band
    # (growth-config default 3.44/4.9/6.36) rather than a measured per-partner census. This is an
    # explicit template assumption — NOT a counted census — and the resulting SAM/TAM proximity to
    # Grab is template-driven, not measured. Replace with a real per-partner census when one exists.
    # (No override entry => falls through to the default M band, which is exactly the template.)
}
# LB-172: Partners without a ridehailing/super-app counterparty skip Journey GMV
# and Partner Platform Rev rungs (no platform take to collect).
PARTNER_NO_PLATFORM_REV = {"french-polynesia"}
# ── Greenfield WIDTH lever — per-partner LOCAL census (golden rule #7: match growth.py) ──
# growth.py auto-discovers finance/recal/greenfield-census/<partner>.json — an ID-based census of
# each partner's OWN sourced-vs-addressable corridors, computed locally from the corridor graph —
# instead of the Grab-global template (4.9). This second cost engine now reads the SAME per-partner
# local census so both engines tell one story and NO sheet needs a hand-patched greenfield cell.
# Precedence:
#   1) --greenfield off / purpose-built captive network  → 1.0 (null-beats-guess)
#   2) explicit PARTNER_MULTIPLIER_OVERRIDES entry        → deliberate cap / reconciliation
#   3) per-partner LOCAL census file                      → ID-based local calculation (matches growth.py)
#   4) else                                               → labelled global template band
GREENFIELD_FORCE_OFF = {"french-polynesia", "bolt-rebase", "saudi-pif", "red-sea-global"}
# --greenfield-json <path>: explicit census override (mirrors growth.py). For a COUNTRY-SPECIFIC
# proposal/sheet, pass the per-partner-COUNTRY census (finance/recal/greenfield-census/<partner>-<country>.json)
# so both engines use the same country-scoped width. Falls back to the whole-partner <partner>.json.
_census_path = arg("--greenfield-json", None) or os.path.join(HERE, "recal", "greenfield-census", f"{PARTNER}.json")
_override = PARTNER_MULTIPLIER_OVERRIDES.get(PARTNER, {})
if arg("--greenfield", None) == "off" or PARTNER in GREENFIELD_FORCE_OFF:
    M["greenfield_corridor_factor"] = {"low":1.0,"mid":1.0,"high":1.0,
        "_doc":"Greenfield width disabled (purpose-built / captive network or --greenfield off)."}
elif _override:
    for _k, _v in _override.items():
        M[_k] = {**M[_k], **_v}
elif os.path.exists(_census_path):
    _cd = json.load(open(_census_path))
    _mode = _cd.get("mode")
    if _mode in ("off", "census_empty"):
        M["greenfield_corridor_factor"] = {"low":1.0,"mid":1.0,"high":1.0,
            "_doc":f"Per-partner local census ({_mode}): no addressable greenfield width beyond the sourced floor."}
    else:
        _fac = _cd["derived_greenfield_factor"]["headline_tier1_plus_tier2"]
        M["greenfield_corridor_factor"] = {"low":float(_fac["low"]),"mid":float(_fac["mid"]),"high":float(_fac["high"]),
            "_doc":("Per-partner LOCAL census (finance/recal/greenfield-census/%s.json): %s sourced / %s greenfield corridors (ratio %s) — ID-based, matches growth.py."
                    % (PARTNER, _cd.get("n_sourced"), _cd.get("n_greenfield_headline"), _cd.get("count_ratio")))}
# LB-254: mirror growth.py mature_capture() exactly — contested ramp (0.15/0.25/0.40) only when
# eff_capture is below the config band; hospitality/captive-blended floors (~0.49–0.55) must use
# max(band, eff_capture) so SAM stays above SOM network (induced × mature > floor capture).
_CAPTIVE_CEILING = 0.95
def _mature_capture_band(eff_capture):
    if eff_capture and eff_capture >= 0.5:
        return {"low": eff_capture, "mid": eff_capture,
                "high": min(_CAPTIVE_CEILING, eff_capture + 0.05)}
    return {b: max(M["mature_capture_rate"][b], eff_capture or 0.0) for b in ("low", "mid", "high")}
if EFF_CAPTURE is not None:
    M["mature_capture_rate"] = _mature_capture_band(EFF_CAPTURE)
SKIP_PLATFORM_REV = (PARTNER in PARTNER_NO_PLATFORM_REV)
def band3(node): return node["low"],node["mid"],node["high"]
# multiplier table
mr=3
sc(ws4,f"A{mr}","Growth multipliers (bands: low / MID / high)",font=H2,fill=f_steel); ws4.merge_cells(f"A{mr}:E{mr}"); mr+=1
for i,h in enumerate(["Multiplier","Low","MID","High","What it means"]): sc(ws4,f"{get_column_letter(i+1)}{mr}",h,font=HDR,fill=f_steel,align=ctr,bd=True)
mr+=1
mult_named={}
def mult(label,node,key,note,fmt=NUM2):
    global mr
    lo,mi,hi=band3(node)
    sc(ws4,f"A{mr}",label,bd=True,font=BOLD)
    sc(ws4,f"B{mr}",lo,fill=f_input,fmt=fmt,align=ctr,bd=True)
    sc(ws4,f"C{mr}",mi,fill=f_input,fmt=fmt,align=ctr,bd=True)
    sc(ws4,f"D{mr}",hi,fill=f_input,fmt=fmt,align=ctr,bd=True)
    sc(ws4,f"E{mr}",note,align=wrap,bd=True,font=SMALL)
    mult_named[key]=(f"$B${mr}",f"$C${mr}",f"$D${mr}"); mr+=1
mult("Induced demand (k)",M["induced_demand"],"ind","Faster/quieter product grows the crossing market beyond today's trips.")
mult("Mature capture rate (c)",M["mature_capture_rate"],"cap","Navier's corridor share once it is the established premium default (vs 10% floor).",PCT)
mult("Journey GMV multiple (m)",M["journey_gmv_multiple"],"gmv","Whole island-journey wallet (transport+food+stay+experiences) as a multiple of the boat fare.")
_green_note = (M.get("greenfield_corridor_factor", {}).get("_doc")
               or "Width lever: addressable network vs the sourced subset (ID-based census).")
mult("Greenfield network factor (g)",M["greenfield_corridor_factor"],"green",_green_note)
# platform take (scalar)
take_r=mr
sc(ws4,f"A{mr}","Platform take rate",bd=True,font=BOLD); sc(ws4,f"C{mr}",v(M["platform_take_rate"]),fill=f_input,fmt=PCT,align=ctr,bd=True)
sc(ws4,f"E{mr}","Super-app blended commission on journey GMV (derived line only).",align=wrap,bd=True,font=SMALL)
take_ref=f"$C${take_r}"; mr+=2
som_cap_r=mr
# LB-254 / exact-parity guard: derive the true sourced transport-spend pool independently
# from grounded corridor demand × comparable fare. Do not recover it from aggregate.py's
# display-rounded effective_capture; that creates small but real model↔sheet ladder drift.
_pool_bucket = "Forward SAM" if ALL_FWD else "Grounded"
_pool_expr = (f"ROUND(SUMPRODUCT(--('Corridor economics'!${FLOORcol}${DATA0}:${FLOORcol}${LASTROW}=\"{_pool_bucket}\"),"
              f"'Corridor economics'!${CL('Demand pool 1-way/yr')}${DATA0}:${CL('Demand pool 1-way/yr')}${LASTROW},"
              f"'Corridor economics'!${CL('Premium fare $/pax')}${DATA0}:${CL('Premium fare $/pax')}${LASTROW}),0)")
# Capture that ACTUALLY built the floor (floor/pool) — captive ~0.90, contested ~0.10,
# blended for mixed. Keep the aggregate value only for narrative/band classification.
_som_cap_val = EFF_CAPTURE if EFF_CAPTURE else v(LAD["som_capture_rate"])
sc(ws4,f"A{mr}","SOM capture rate (floor)",bd=True,font=BOLD); sc(ws4,f"C{mr}",f"=som_floor_rev/({_pool_expr})",fill=f_calc,fmt=PCT,align=ctr,bd=True)
sc(ws4,f"E{mr}",("Captive sole-operator capture (~%.0f%%) that builds the published floor; M_today = floor \u00f7 this = the true corridor spend pool." % (_som_cap_val*100)
                 if IS_CAPTIVE and EFF_CAPTURE else
                 "New-entrant ~%.0f%% share of today's pool (= the published floor)." % (_som_cap_val*100)),align=wrap,bd=True,font=SMALL)
somcap_ref=f"$C${som_cap_r}"; mr+=2

# anchor + ladder
sc(ws4,f"A{mr}","The ladder (live off the corridor floor)",font=H2,fill=f_steel); ws4.merge_cells(f"A{mr}:E{mr}"); mr+=1
for i,h in enumerate(["Rung","Low","MID","High","Definition"]): sc(ws4,f"{get_column_letter(i+1)}{mr}",h,font=HDR,fill=f_steel,align=ctr,bd=True)
mr+=1
def lad(label,lo,mi,hi,note,fmt=USD):
    global mr
    sc(ws4,f"A{mr}",label,bd=True,font=BOLD)
    sc(ws4,f"B{mr}",lo,fill=f_calc,fmt=fmt,align=ctr,bd=True)
    sc(ws4,f"C{mr}",mi,fill=f_calc,fmt=fmt,align=ctr,bd=True)
    sc(ws4,f"D{mr}",hi,fill=f_calc,fmt=fmt,align=ctr,bd=True)
    sc(ws4,f"E{mr}",note,align=wrap,bd=True,font=SMALL); mr+=1
ind_lo,ind_mi,ind_hi=mult_named["ind"]; cap_lo,cap_mi,cap_hi=mult_named["cap"]
gmv_lo,gmv_mi,gmv_hi=mult_named["gmv"]; gr_lo,gr_mi,gr_hi=mult_named["green"]
# SOM floor (single) + M_today anchor
sc(ws4,f"A{mr}",("FORWARD-SAM anchor \u2014 Navier transport rev/yr (2030-dated; NOT a grounded floor)" if ALL_FWD else "SOM floor \u2014 Navier transport rev/yr (PUBLISHED)"),bd=True,font=BOLD)
sc(ws4,f"C{mr}","=som_floor_rev",fill=f_calc,fmt=USD,align=ctr,bd=True)
sc(ws4,f"E{mr}",("Live sum of forward-SAM corridor market-rev (10% capture, 2030-dated destination-cap demand). Held OUT of any near-term grounded number." if ALL_FWD else "Live sum of grounded-floor corridor market-rev only (Floor bucket = Grounded; cascade-estimated rows held out). Matches growth.py _headline_anchor and deck TAM."),align=wrap,bd=True,font=SMALL)
som_floor_ref=f"$C${mr}"; mr+=1
sc(ws4,f"A{mr}","M_today \u2014 model comparable-fare transport pool on sourced corridors",bd=True,font=BOLD)
sc(ws4,f"C{mr}",f"={_pool_expr}",fill=f_calc,fmt=USD,align=ctr,bd=True)
sc(ws4,f"E{mr}","Direct sum of sourced one-way demand \u00d7 the disclosed model comparable fare on grounded corridors; includes any labelled premium re-fare applied above and is independent of display-rounded capture.",align=wrap,bd=True,font=SMALL)
Mt=f"$C${mr}"; mr+=2
lad(("SOM full network (~%.0f%% capture, today, +greenfield)" % (_som_cap_val*100)),
    f"={Mt}*{somcap_ref}*{gr_lo}",f"={Mt}*{somcap_ref}*{gr_mi}",f"={Mt}*{somcap_ref}*{gr_hi}",
    ("Floor posture (~%.0f%% captive capture, today's demand) extended across the whole mapped network." % (_som_cap_val*100)))
lad("SAM \u2014 sourced corridors only (depth, no greenfield)",
    f"={Mt}*{ind_lo}*{cap_lo}",f"={Mt}*{ind_mi}*{cap_mi}",f"={Mt}*{ind_hi}*{cap_hi}",
    "Matured network on the sourced corridors alone: induced demand \u00d7 mature capture.")
lad("SAM \u2014 Navier transport rev @ full network (HEADLINE)",
    f"={Mt}*{ind_lo}*{cap_lo}*{gr_lo}",f"={Mt}*{ind_mi}*{cap_mi}*{gr_mi}",f"={Mt}*{ind_hi}*{cap_hi}*{gr_hi}",
    "Matured across the full addressable network: induced \u00d7 mature capture \u00d7 greenfield width.")
lad("Marine mobility TAM \u2014 induced marine-transfer market (LB-110)",
    f"={Mt}*{ind_lo}*{gr_lo}",f"={Mt}*{ind_mi}*{gr_mi}",f"={Mt}*{ind_hi}*{gr_hi}",
    "SAM divided by leading-operator capture (= M_today \u00d7 induced \u00d7 greenfield). The full inducible water-transfer wallet at network maturity, before any capture constraint.")
if not SKIP_PLATFORM_REV:
    lad("Journey GMV \u2014 induced crossing \u00d7 multiple (was TAM journey GMV)",
        f"={Mt}*{ind_lo}*{gmv_lo}*{gr_lo}",f"={Mt}*{ind_mi}*{gmv_mi}*{gr_mi}",f"={Mt}*{ind_hi}*{gmv_hi}*{gr_hi}",
        "Whole-journey spend (transport+food+stay+experiences) across the induced crossing market on the full network. Marine TAM \u00d7 journey GMV multiple.")
    lad("Journey GMV routed through Navier network",
        f"={Mt}*{ind_lo}*{cap_lo}*{gmv_lo}*{gr_lo}",f"={Mt}*{ind_mi}*{cap_mi}*{gmv_mi}*{gr_mi}",f"={Mt}*{ind_hi}*{cap_hi}*{gmv_hi}*{gr_hi}",
        "Journey GMV on Navier-carried trips only.")
    lad("Partner platform revenue on Navier (LB-113 \u2014 on-Navier subset, not 18% of full Journey GMV)",
        f"={Mt}*{ind_lo}*{cap_lo}*{gmv_lo}*{gr_lo}*{take_ref}",f"={Mt}*{ind_mi}*{cap_mi}*{gmv_mi}*{gr_mi}*{take_ref}",f"={Mt}*{ind_hi}*{cap_hi}*{gmv_hi}*{gr_hi}*{take_ref}",
        "The super-app's own commission \u2014 the number a platform partner cares about.")
mr+=1
sc(ws4,f"A{mr}","Note",font=BOLD); sc(ws4,f"B{mr}","All rungs are one multiplication chain off M_today, which is live off the corridor floor. Change any input (a fare, a multiplier, the scenario) and the whole ladder moves. Bands (low/MID/high) are never single points; greenfield census is per-partner.",align=wrap,font=SMALL); ws4.merge_cells(f"B{mr}:E{mr}")

# ------------------------------------------------------------ register names
for name,ref in named.items():
    try:
        wb.defined_names.add(DefinedName(name, attr_text=ref))
    except Exception as e:
        print("name fail",name,e)

# roadmap note on corridor tab
if roadmap:
    rn=(FSROW or UPROW or BUT)+2
    sc(ws,f"{CL('Market')}{rn}","Held for Quanta-LR (>70nm, roadmap H2 2026+):",font=Font(italic=True,bold=True,size=9),align=Alignment(vertical="center"))
    for j,rm in enumerate(roadmap):
        sc(ws,f"{CL('Corridor')}{rn+j}",f"{rm['corridor']} ({rm['nm']}nm)",font=Font(size=9,italic=True),align=wrap)

wb.save(OUT)
print("wrote", OUT)
print("engine rows:", len(rows), "| roadmap:", len(roadmap), "| countries:", len(used_countries))
print("named ranges:", len(named))
