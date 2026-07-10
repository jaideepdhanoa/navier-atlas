#!/usr/bin/env python3
"""Master tracker: partner index + global unit-economics on unique geometry.

Tabs:
  - Partner unit-economics — one row per partner sheet (gross rollup, not de-duped)
  - Assumptions / Country opex / Corridor economics / Global TAM
    Corridor economics = canonical de-duped pier-pair list (one row per crossing);
    Global TAM ladder formulas hang off its grounded-floor named ranges.

Reads rollups from finance/recal/agg-<partner>.json (no hand-transcription).
"""
from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import tempfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
try:
    from partner_keys import engine_partner, sheet_partner
except ImportError:
    from finance.partner_keys import engine_partner, sheet_partner  # type: ignore

HERE = os.path.dirname(os.path.abspath(__file__))
MODEL = os.path.join(HERE, "model")
DEFAULT_AGG_DIR = os.path.join(HERE, "recal")
DEFAULT_OUT = os.path.join(HERE, "_master-unit-econ.xlsx")
SHEET_IDS_PATH = os.path.join(HERE, "PARTNER-SHEET-IDS.json")
TRANSPARENT_BUILDER = os.path.join(HERE, "build_transparent_sheet.py")
AGG_GLOBAL = os.path.join(DEFAULT_AGG_DIR, "agg-global.json")
AGG_UNIQUE = os.path.join(DEFAULT_AGG_DIR, "agg-unique-global.json")
GROWTH_UNIQUE = os.path.join(DEFAULT_AGG_DIR, "growth-unique-global.json")

GLOBAL_ECON_SHEETS = ("Assumptions", "Country opex", "Corridor economics", "Global TAM")

# Stable row order; any slug in PARTNER-SHEET-IDS.json not listed here appends alphabetically.
PARTNER_ORDER = [
    "grab", "careem", "bolt", "yango", "didi", "uber",
    "rapido", "ola", "noon",
    "jih-global", "qatar", "saudi-pif", "red-sea-global",
    "constance", "four-seasons",
]

PARTNER_DISPLAY: dict[str, str] = {
    "grab": "Grab",
    "careem": "Careem",
    "bolt": "Bolt",
    "yango": "Yango",
    "didi": "DiDi (Mexico + Brazil calibration)",
    "uber": "Uber",
    "rapido": "Rapido (India)",
    "ola": "Ola (India)",
    "noon": "Noon (UAE)",
    "jih-global": "JIH Global (Maldives)",
    "qatar": "Qatar",
    "saudi-pif": "Saudi / Red Sea (PIF)",
    "red-sea-global": "Red Sea Global",
    "constance": "Constance (Maldives)",
    "four-seasons": "Four Seasons (Maldives)",
}

PARTNER_NOTES: dict[str, str] = {
    "grab": "SE-Asia ridehail/island network. Primary anchor — 0% future-dated demand.",
    "careem": "UAE luxury water-served transfers. Mostly estimated; corridors city-level (route-pin pending).",
    "saudi-pif": "Low-confidence / largely 2030-dated. Greenfield OFF. Treat as forward.",
    "red-sea-global": "Resort-operator corridors. Small, grounded subset.",
    "jih-global": "Maldives 43 corridors geometry-bound; network-sum captive fleet.",
    "qatar": "Doha 4/4 in-range corridors bound. Banana Island captive carries floor.",
    "bolt": "MENA grounded floor + Med/Baltic aspirational tail (Bucket C bound).",
    "yango": "MENA + Africa grounded + CIS/Caspian/Africa aspirational tail (Bucket C bound).",
    "didi": "Mexico + Brazil calibration — two exact Caribbean anchors and four exact Rio lines; Pacific and Colombia demand held null.",
    "constance": "Captive resort-transfer; network-sum captive fleet. Greenfield OFF.",
    "four-seasons": "Captive resort-transfer; network-sum captive fleet. Greenfield OFF.",
    "uber": "Global mobility rollup; MENA + Med + Hawaii + LatAm scoped markets.",
    "rapido": "PR #58 India — Mumbai/Goa/Kerala/Andaman sealed spine; ridehail archetype.",
    "ola": "PR #58 India — same sealed spine as Rapido; ridehail archetype.",
    "noon": "PR #58 UAE — 12 sealed super_app corridors; Careem-style demand ladder.",
}

NAVY = PatternFill("solid", fgColor="1F3A5F")
STEEL = PatternFill("solid", fgColor="2E5984")
ZEBRA = PatternFill("solid", fgColor="EEF3F8")
TOTF = PatternFill("solid", fgColor="DCE6F1")
WHT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
SMALL = Font(size=9, color="555555")
thin = Side(style="thin", color="BBBBBB")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
USD = "$#,##0"
NUM = "#,##0"
NUM1 = "#,##0.0"
PCT = "0.0%"
ctr = Alignment(horizontal="center", vertical="center")
wrap = Alignment(wrap_text=True, vertical="top")


def load_registry_rows() -> list[tuple[str, str, str, str]]:
    """(display, sheet-slug, google-sheet-id, note) from PARTNER-SHEET-IDS.json."""
    reg = json.load(open(SHEET_IDS_PATH))
    slugs = [k for k in reg if not k.startswith("_")]
    order = {s: i for i, s in enumerate(PARTNER_ORDER)}
    slugs.sort(key=lambda s: (order.get(s, 999), s))
    rows: list[tuple[str, str, str, str]] = []
    for slug in slugs:
        sid = reg[slug]
        if not sid or str(sid).startswith("_"):
            continue
        disp = PARTNER_DISPLAY.get(slug, slug.replace("-", " ").title())
        note = PARTNER_NOTES.get(slug, "Unit-economics transparent sheet — see agg rollup when present.")
        rows.append((disp, slug, sid, note))
    return rows


def agg_path(agg_dir: str, sheet_slug: str):
    """Resolve agg JSON for a sheet slug (handles saudi-pif ↔ saudi-redsea-pif)."""
    candidates = [
        os.path.join(agg_dir, f"agg-{sheet_slug}.json"),
        os.path.join(agg_dir, f"agg-{engine_partner(sheet_slug)}.json"),
    ]
    seen: set[str] = set()
    for path in candidates:
        if path in seen:
            continue
        seen.add(path)
        if os.path.isfile(path):
            return path
    return None


def rollup_from_agg(path: str) -> dict:
    j = json.load(open(path))
    d = j["rollup"]
    gf, eu, et = d.get("grounded_floor", {}), d.get("estimated_upside", {}), d.get("estimated_total", {})
    rows = j.get("rows", [])
    by_status: dict[str, int] = {}
    for r in rows:
        by_status[r.get("status", "?")] = by_status.get(r.get("status", "?"), 0) + 1
    g_rows = by_status.get("grounded", d.get("n_grounded", 0))
    e_rows = by_status.get("estimated", d.get("n_estimated", 0))
    gfleet = gf.get("fleet", 0)
    grev = gf.get("market_rev_yr", 0)
    efleet = eu.get("fleet", 0)
    erev = eu.get("market_rev_yr", 0)
    tfleet = et.get("fleet", 0)
    trev = et.get("market_rev_yr", 0)
    scope_only = (tfleet or 0) < (gfleet or 0)
    if scope_only:
        efleet, erev = None, None
        tfleet, trev = gfleet, grev
    return dict(
        n=d.get("n_corridors_total", 0), g=g_rows, e=e_rows,
        gfleet=gfleet, grev=grev, efleet=efleet, erev=erev,
        tfleet=tfleet, trev=trev, scope_only=scope_only,
        agg_file=path,
    )


def empty_rollup(reason: str) -> dict:
    return dict(
        n=None, g="—", e="—",
        gfleet=None, grev=None, efleet=None, erev=None,
        tfleet=None, trev=None, scope_only=False,
        agg_file=None, missing=reason,
    )


def rollup(sheet_slug: str, agg_dir: str) -> dict:
    path = agg_path(agg_dir, sheet_slug)
    if not path:
        return empty_rollup(f"no agg-{sheet_slug}.json in {agg_dir}")
    return rollup_from_agg(path)


def ensure_global_aggs(agg_dir: str) -> None:
    """Build agg-global + agg-unique-global + growth-unique-global if missing or stale."""
    global_path = os.path.join(agg_dir, "agg-global.json")
    unique_path = os.path.join(agg_dir, "agg-unique-global.json")
    growth_path = os.path.join(agg_dir, "growth-unique-global.json")
    agg_py = os.path.join(MODEL, "aggregate.py")
    growth_py = os.path.join(MODEL, "growth.py")

    if not os.path.isfile(global_path):
        subprocess.run(
            ["python3", agg_py, "--partner", "global", "--json", global_path],
            check=True,
        )
    subprocess.run(
        ["python3", agg_py, "--partner", "global", "--dedup", "unique", "--json", unique_path],
        check=True,
    )
    subprocess.run(
        ["python3", growth_py, "--agg", unique_path, "--partner", "global-unique", "--json", growth_path],
        check=True,
    )


def build_global_econ_workbook(out_path: str, agg_dir: str) -> None:
    unique_agg = os.path.join(agg_dir, "agg-unique-global.json")
    cmd = [
        "python3", TRANSPARENT_BUILDER,
        "--partner", "global",
        "--dedup", "unique",
        "--agg", unique_agg,
        "--out", out_path,
        "--skip-readme",
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        raise RuntimeError(f"global transparent sheet failed:\n{proc.stderr}\n{proc.stdout}")
    print(proc.stdout.strip())


def copy_worksheet(source_ws, target_wb, title: str | None = None):
    """Copy a worksheet from another workbook into target_wb."""
    from copy import copy

    name = title or source_ws.title
    if name in target_wb.sheetnames:
        del target_wb[name]
    tgt = target_wb.create_sheet(name)
    for row in source_ws.iter_rows():
        for cell in row:
            tc = tgt[cell.coordinate]
            tc.value = cell.value
            if cell.has_style:
                tc.font = copy(cell.font)
                tc.border = copy(cell.border)
                tc.fill = copy(cell.fill)
                tc.number_format = cell.number_format
                tc.protection = copy(cell.protection)
                tc.alignment = copy(cell.alignment)
    for col, dim in source_ws.column_dimensions.items():
        tgt.column_dimensions[col].width = dim.width
        tgt.column_dimensions[col].hidden = dim.hidden
    for row, dim in source_ws.row_dimensions.items():
        tgt.row_dimensions[row].height = dim.height
        tgt.row_dimensions[row].hidden = dim.hidden
    tgt.freeze_panes = source_ws.freeze_panes
    if source_ws.auto_filter.ref:
        tgt.auto_filter.ref = source_ws.auto_filter.ref
    for mref in source_ws.merged_cells.ranges:
        tgt.merge_cells(str(mref))
    return tgt


def merge_global_econ_tabs(master_wb: openpyxl.Workbook, global_xlsx: str) -> None:
    src = openpyxl.load_workbook(global_xlsx, data_only=False)
    for name in GLOBAL_ECON_SHEETS:
        if name not in src.sheetnames:
            alt = "Market sizing" if name == "Global TAM" else None
            if alt and alt in src.sheetnames:
                copy_worksheet(src[alt], master_wb, "Global TAM")
            else:
                print(f"warning: missing sheet {name} in {global_xlsx}", file=sys.stderr)
            continue
        copy_worksheet(src[name], master_wb, name)
    # copy defined names (global econ engine)
    for defn in src.defined_names.values():
        master_wb.defined_names.add(defn)


def _put(ws, ref, val, font=None, fill=None, fmt=None, align=None, bd=True):
    c = ws[ref]
    c.value = val
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if bd:
        c.border = BORD


def append_growth_summary(ws) -> None:
    """Append engine-verified TAM headline block below the formula ladder on Global TAM tab."""
    if not os.path.isfile(GROWTH_UNIQUE):
        return
    g = json.load(open(GROWTH_UNIQUE))
    anchor = g.get("grounded") or {}
    if not anchor.get("SAM_navier_transport_rev_yr"):
        return
    r = ws.max_row + 3
    _put(ws, f"A{r}", "Engine cross-check (growth-unique-global.json — unique geometry)", BOLD, TOTF, align=wrap)
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    lines = [
        ("M_today pool", anchor.get("M_today_transport_spend_yr")),
        ("SOM floor (published)", anchor.get("SOM_floor_navier_transport_rev_yr")),
        ("SAM @ full network (MID)", (anchor.get("SAM_navier_transport_rev_yr") or {}).get("mid")),
        ("TAM journey GMV (MID)", (anchor.get("TAM_journey_gmv_yr") or {}).get("mid")),
        ("Platform rev on Navier (MID)", (anchor.get("partner_platform_rev_yr") or {}).get("mid")),
    ]
    for label, val in lines:
        _put(ws, f"A{r}", label, SMALL, None, align=wrap)
        _put(ws, f"C{r}", val, BOLD, TOTF, USD, ctr)
        r += 1


def build_partner_index_sheet(wb: openpyxl.Workbook, agg_dir: str) -> int:
    ws = wb.active
    ws.title = "Partner unit-economics"

    ws.merge_cells("A1:J1")
    _put(ws, "A1", "Navier — Partner Corridor Unit-Economics  ·  master tracker",
         Font(color="FFFFFF", bold=True, size=14), NAVY,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:J2")
    _put(ws, "A2",
         "Each partner sheet is fully formula-driven and standalone (Assumptions · Corridor economics · Market sizing). "
         "Numbers below are the engine rollup; MID scenario. Open a sheet to trace any cell. "
         "Portfolio total is gross (partner scopes may overlap). "
         "Global unit-economics (de-duped Corridor economics tab) + Global TAM are on the tabs to the right.",
         SMALL, None, align=wrap, bd=False)
    ws.row_dimensions[2].height = 36

    hdr = ["Partner", "Open sheet", "Corridors (g / e)", "Grounded fleet", "Grounded mkt rev/yr",
           "Est. upside fleet", "Est. upside rev/yr", "Total fleet", "Total rev/yr", "Notes"]
    for i, h in enumerate(hdr):
        _put(ws, f"{chr(65 + i)}3", h, WHT, STEEL,
             align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[3].height = 34

    r = 4
    missing: list[str] = []
    for disp, pk, sid, note in load_registry_rows():
        ro = rollup(pk, agg_dir)
        if ro.get("missing"):
            missing.append(pk)
            note = f"{note} [{ro['missing']}]"
        fill = ZEBRA if (r % 2 == 0) else None
        _put(ws, f"A{r}", disp, BOLD, fill)
        _put(ws, f"B{r}", f'=HYPERLINK("https://docs.google.com/spreadsheets/d/{sid}/edit","open \u2197")',
             Font(color="1155CC", underline="single"), fill, align=ctr)
        _put(ws, f"C{r}", f'{ro["g"]} / {ro["e"]}', None, fill, align=ctr)
        _put(ws, f"D{r}", ro["gfleet"], None, fill, align=ctr)
        _put(ws, f"E{r}", ro["grev"], None, fill, USD)
        _put(ws, f"F{r}", ro["efleet"], None, fill, align=ctr)
        _put(ws, f"G{r}", ro["erev"], None, fill, USD)
        _put(ws, f"H{r}", ro["tfleet"], None, fill, align=ctr)
        _put(ws, f"I{r}", ro["trev"], None, fill, USD)
        _put(ws, f"J{r}", note, SMALL, fill, align=wrap)
        ws.row_dimensions[r].height = 42
        r += 1

    _put(ws, f"A{r}", "PORTFOLIO (gross)", BOLD, TOTF)
    _put(ws, f"B{r}", "", None, TOTF)
    _put(ws, f"C{r}", "", None, TOTF)
    for col in ["D", "E", "F", "G", "H", "I"]:
        fmt = USD if col in ("E", "G", "I") else None
        _put(ws, f"{col}{r}", f"=SUM({col}4:{col}{r - 1})", BOLD, TOTF, fmt,
             align=(ctr if col in ("D", "F", "H") else None))
    _put(ws, f"J{r}", "Gross sum across partner scopes; not de-duplicated.", SMALL, TOTF, align=wrap)
    ws.row_dimensions[r].height = 30

    widths = {"A": 22, "B": 12, "C": 15, "D": 13, "E": 18, "F": 14, "G": 18, "H": 12, "I": 18, "J": 46}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A4"

    if missing:
        print(f"warning: missing agg for: {', '.join(missing)}", file=sys.stderr)
    return r


def build_master(out_path: str, agg_dir: str = DEFAULT_AGG_DIR, *, skip_global: bool = False) -> int:
    ensure_global_aggs(agg_dir)

    wb = openpyxl.Workbook()
    last_row = build_partner_index_sheet(wb, agg_dir)

    if not skip_global:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            global_path = tmp.name
        try:
            build_global_econ_workbook(global_path, agg_dir)
            merge_global_econ_tabs(wb, global_path)
            if "Global TAM" in wb.sheetnames:
                append_growth_summary(wb["Global TAM"])
        finally:
            if os.path.isfile(global_path):
                os.unlink(global_path)

    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    wb.save(out_path)
    print(f"wrote {out_path} ; partner rows: {last_row} ; tabs: {wb.sheetnames}")
    return last_row


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=DEFAULT_OUT)
    ap.add_argument("--agg-dir", default=DEFAULT_AGG_DIR)
    ap.add_argument("--skip-global", action="store_true", help="partner index only (legacy)")
    args = ap.parse_args()
    build_master(args.out, args.agg_dir, skip_global=args.skip_global)


if __name__ == "__main__":
    main()