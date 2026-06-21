#!/usr/bin/env python3
"""Build the partner tracking sheet — all partners, markets, Atlas + financials links.

Reads live partner-pitch JSON (+ _draft), writes finance/_partner-tracker.xlsx, and
optionally uploads in-place to the Google Sheet registered as _partner_tracker in
PARTNER-SHEET-IDS.json.

Tabs:
  - Markets   — one row per partner market (primary ops view)
  - Partners  — one row per partner (rollup)
  - Meta      — generation stamp + counts
"""
from __future__ import annotations

import argparse
import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from drive_upload import replace_spreadsheet
except ImportError:
    from finance.drive_upload import replace_spreadsheet  # type: ignore

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
PARTNERS_DIR = ROOT / "partner-pitch" / "partners"
DRAFT_DIR = PARTNERS_DIR / "_draft"
DEFAULT_OUT = HERE / "_partner-tracker.xlsx"
SHEET_IDS_PATH = HERE / "PARTNER-SHEET-IDS.json"
URL_MAP_PATH = HERE / "economics_url_map.json"
ATLAS_BASE = os.environ.get("ATLAS_BASE_URL", "https://navier-atlas.vercel.app").rstrip("/")
DRIVE_URL_FMT = "https://docs.google.com/spreadsheets/d/{sid}/edit"

NAVY = PatternFill("solid", fgColor="1F3A5F")
ZEBRA = PatternFill("solid", fgColor="EEF3F8")
WHT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
LINK = Font(color="0563C1", underline="single")
SMALL = Font(size=9, color="555555")
thin = Side(style="thin", color="BBBBBB")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
ctr = Alignment(horizontal="center", vertical="center")

MARKET_COLS = [
    "Partner",
    "Partner ID",
    "Market",
    "Market slug",
    "Category",
    "Archetype",
    "Region",
    "Layout",
    "Atlas",
    "Financials",
    "Deck",
    "Status",
    "Economics",
    "Tier",
    "Anchor cities",
    "Markets count",
    "Source",
    "Updated",
]

PARTNER_COLS = [
    "Partner",
    "Partner ID",
    "Category",
    "Archetype",
    "Region",
    "Layout",
    "Markets",
    "Atlas",
    "Financials",
    "Deck",
    "Status",
    "Economics",
    "Proposal",
    "Tier",
    "Source",
    "Updated",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def load_json(path: Path) -> Any:
    return json.loads(path.read_text())


def sheet_url(sid: str) -> str:
    return DRIVE_URL_FMT.format(sid=sid)


def atlas_partner_url(partner_id: str) -> str:
    return f"{ATLAS_BASE}/{partner_id}"


def atlas_market_url(partner_id: str, slug: str) -> str:
    return f"{ATLAS_BASE}/{partner_id}/{slug}"


def financials_url(
    partner_id: str,
    doc: dict[str, Any],
    *,
    economics_map: dict[str, str],
    sheet_ids: dict[str, str],
) -> str | None:
    url = doc.get("economics_url")
    if isinstance(url, str) and url.strip():
        return url.strip()
    mapped = economics_map.get(partner_id)
    if mapped:
        return mapped
    sid = sheet_ids.get(partner_id)
    if sid and not str(sid).startswith("_"):
        return sheet_url(sid)
    growth = doc.get("growth_case") or {}
    if isinstance(growth, dict):
        gurl = growth.get("economics_url")
        if isinstance(gurl, str) and gurl.strip():
            return gurl.strip()
    return None


def status_label(doc: dict[str, Any]) -> str:
    parts: list[str] = []
    ps = doc.get("proposal_status")
    if isinstance(ps, str) and ps.strip():
        if "seal" in ps.lower() or "complete" in ps.lower():
            parts.append("Sealed")
        elif "pending" in ps.lower():
            parts.append("Seal pending")
        else:
            parts.append(ps.replace("_", " "))
    es = doc.get("economics_status")
    if es == "economics_pending":
        parts.append("Economics pending")
    elif isinstance(es, str) and es.strip():
        parts.append(es.replace("_", " "))
    tier = doc.get("tier")
    if tier == "review":
        parts.append("Review tier")
    if not parts:
        return "Live"
    return " · ".join(parts)


def iter_partner_files() -> list[tuple[Path, str]]:
    """(path, source_tag) — pitch partners first, then drafts not already present."""
    seen: set[str] = set()
    out: list[tuple[Path, str]] = []
    for path in sorted(PARTNERS_DIR.glob("*.json")):
        if path.name.startswith("_"):
            continue
        out.append((path, "partner-pitch"))
        seen.add(path.stem)
    for path in sorted(DRAFT_DIR.glob("*.json")):
        if path.stem in seen:
            continue
        out.append((path, "draft"))
    return out


def collect_records(
    *,
    economics_map: dict[str, str],
    sheet_ids: dict[str, str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    partner_rows: list[dict[str, Any]] = []
    market_rows: list[dict[str, Any]] = []
    stamp = utc_now()

    for path, source in iter_partner_files():
        doc = load_json(path)
        pid = doc.get("partner_id") or path.stem
        display = doc.get("display") or pid.replace("-", " ").title()
        layout = doc.get("layout") or "single"
        category = doc.get("category") or ""
        archetype = doc.get("archetype") or ""
        region = doc.get("region") or ""
        fin = financials_url(pid, doc, economics_map=economics_map, sheet_ids=sheet_ids)
        stat = status_label(doc)
        atlas = atlas_partner_url(pid)
        markets = doc.get("markets") or []

        partner_rows.append({
            "Partner": display,
            "Partner ID": pid,
            "Category": category,
            "Archetype": archetype,
            "Region": region,
            "Layout": layout,
            "Markets": len(markets) if layout == "hub" and markets else (1 if layout == "single" else 0),
            "Atlas": atlas,
            "Financials": fin or "",
            "Deck": "",
            "Status": stat,
            "Economics": doc.get("economics_status") or "",
            "Proposal": doc.get("proposal_status") or "",
            "Tier": doc.get("tier") or "",
            "Source": source,
            "Updated": stamp,
        })

        if layout == "hub" and markets:
            for m in markets:
                slug = m.get("slug") or m.get("id") or ""
                anchors = m.get("anchor_cities") or []
                market_rows.append({
                    "Partner": display,
                    "Partner ID": pid,
                    "Market": m.get("label") or slug,
                    "Market slug": slug,
                    "Category": m.get("category") or category,
                    "Archetype": archetype,
                    "Region": m.get("region") or region,
                    "Layout": layout,
                    "Atlas": atlas_market_url(pid, slug) if slug else atlas,
                    "Financials": fin or "",
                    "Deck": "",
                    "Status": m.get("scope_status") or stat,
                    "Economics": doc.get("economics_status") or "",
                    "Tier": doc.get("tier") or "",
                    "Anchor cities": ", ".join(anchors) if anchors else "(brief-only)",
                    "Markets count": len(markets),
                    "Source": source,
                    "Updated": stamp,
                })
        else:
            anchors: list[str] = []
            for phase in doc.get("phases") or []:
                for c in phase.get("cities") or []:
                    if c not in anchors:
                        anchors.append(c)
            market_rows.append({
                "Partner": display,
                "Partner ID": pid,
                "Market": region or display,
                "Market slug": "",
                "Category": category,
                "Archetype": archetype,
                "Region": region,
                "Layout": layout,
                "Atlas": atlas,
                "Financials": fin or "",
                "Deck": "",
                "Status": stat,
                "Economics": doc.get("economics_status") or "",
                "Tier": doc.get("tier") or "",
                "Anchor cities": ", ".join(anchors) if anchors else "",
                "Markets count": 1,
                "Source": source,
                "Updated": stamp,
            })

    partner_rows.sort(key=lambda r: (r["Region"], r["Partner"]))
    market_rows.sort(key=lambda r: (r["Region"], r["Partner"], r["Market"]))
    return partner_rows, market_rows


def _set_link(cell, url: str, label: str | None = None) -> None:
    if not url:
        cell.value = "—"
        return
    cell.value = label or url
    cell.hyperlink = url
    cell.font = LINK


def _write_table(ws, headers: list[str], rows: list[dict[str, Any]], link_cols: set[str]) -> None:
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill = NAVY
        c.font = WHT
        c.alignment = ctr
        c.border = BORD

    for i, row in enumerate(rows, start=2):
        fill = ZEBRA if i % 2 == 0 else PatternFill()
        for col, h in enumerate(headers, start=1):
            val = row.get(h, "")
            cell = ws.cell(i, col, val if h not in link_cols else (val or "—"))
            cell.border = BORD
            cell.alignment = wrap
            if fill != PatternFill():
                cell.fill = fill
            if h in link_cols and val:
                _set_link(cell, val, "Open" if h in ("Atlas", "Financials", "Deck") else val)

    for col, h in enumerate(headers, 1):
        width = 14
        if h in ("Partner", "Market", "Status", "Anchor cities"):
            width = 28
        elif h in ("Atlas", "Financials"):
            width = 12
        elif h in ("Partner ID", "Market slug"):
            width = 22
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def build_workbook(
    out_path: Path,
    *,
    partner_rows: list[dict[str, Any]],
    market_rows: list[dict[str, Any]],
) -> None:
    wb = openpyxl.Workbook()
    ws_markets = wb.active
    ws_markets.title = "Markets"
    _write_table(ws_markets, MARKET_COLS, market_rows, {"Atlas", "Financials", "Deck"})

    ws_partners = wb.create_sheet("Partners")
    _write_table(ws_partners, PARTNER_COLS, partner_rows, {"Atlas", "Financials", "Deck"})

    ws_meta = wb.create_sheet("Meta")
    meta = [
        ("Generated", utc_now()),
        ("Atlas base", ATLAS_BASE),
        ("Partner count", len(partner_rows)),
        ("Market rows", len(market_rows)),
        ("With financials", sum(1 for r in partner_rows if r.get("Financials"))),
        ("Source", "finance/build_partner_tracker.py"),
        ("Registry", str(SHEET_IDS_PATH.relative_to(ROOT))),
    ]
    for i, (k, v) in enumerate(meta, start=1):
        ws_meta.cell(i, 1, k).font = BOLD
        ws_meta.cell(i, 2, v)
    ws_meta.column_dimensions["A"].width = 18
    ws_meta.column_dimensions["B"].width = 48

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def build(out_path: Path | None = None) -> dict[str, Any]:
    out = out_path or DEFAULT_OUT
    reg = load_json(SHEET_IDS_PATH) if SHEET_IDS_PATH.is_file() else {}
    url_doc = load_json(URL_MAP_PATH) if URL_MAP_PATH.is_file() else {}
    economics_map = url_doc.get("economics_url") or {}
    sheet_ids = {k: v for k, v in reg.items() if not k.startswith("_")}

    partner_rows, market_rows = collect_records(
        economics_map=economics_map,
        sheet_ids=sheet_ids,
    )
    build_workbook(out, partner_rows=partner_rows, market_rows=market_rows)
    return {
        "out": str(out),
        "partners": len(partner_rows),
        "markets": len(market_rows),
        "with_financials": sum(1 for r in partner_rows if r.get("Financials")),
        "generated_at": utc_now(),
    }


def upload(out_path: Path, *, dry_run: bool = False) -> dict[str, Any]:
    reg = load_json(SHEET_IDS_PATH)
    tracker_id = reg.get("_partner_tracker")
    if not tracker_id:
        return {"status": "skipped", "reason": "no _partner_tracker in PARTNER-SHEET-IDS.json"}
    return replace_spreadsheet(str(out_path), tracker_id, dry_run=dry_run)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--upload", action="store_true", help="Upload to _partner_tracker Google Sheet")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    summary = build(Path(args.out))
    print(json.dumps(summary, indent=2))

    if args.upload:
        up = upload(Path(args.out), dry_run=args.dry_run)
        print(json.dumps(up, indent=2))
        if up.get("status") == "upload-failed":
            return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())