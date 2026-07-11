#!/usr/bin/env python3
"""Focused regression coverage for the South Korea schedule/occupancy correction."""
from __future__ import annotations

import copy
import importlib.util
import json
import math
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

HERE = Path(__file__).resolve().parent
FINANCE = HERE.parent
MODEL = FINANCE / "model"
FIXTURE = HERE / "fixtures" / "south-korea-schedule-corridors.json"


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


ATOM = load_module("korea_test_atom", MODEL / "atom.py")
AGGREGATE = load_module("korea_test_aggregate", MODEL / "aggregate.py")
CONST = json.loads((MODEL / "vessel-constants.json").read_text())
KOREA = CONST["operating_defaults"]["country_operating_overrides"]["South Korea"]


def corridor(country: str, distance_nm: float) -> dict:
    return {
        "route_id": "test-route",
        "from": "A",
        "to": "B",
        "country": country,
        "distance_nm": distance_nm,
        "archetype": "ridehail",
        "L3_locals": {
            "comparable_fare_usd_pax": 30,
            "demand_ferry_rides_yr": 100000,
            "energy_usd_per_kwh": 0.12,
            "captain_annual_usd": 40000,
            "marina_overhead_annual_usd": 10000,
            "grid_kg_co2_per_kwh": 0.4,
            "weather_uptime_factor": 1.0,
        },
    }


class ModelEngineTests(unittest.TestCase):
    def test_korea_short_route_is_schedule_derived_and_exceeds_old_cap(self):
        scenarios = AGGREGATE.run_scenarios(corridor("South Korea", 1.1), "pioneer_ii")
        mid = scenarios["mid"]
        charge_min = 1.1 / 70 * 45
        cycle_min = 1.1 / 20 * 60 + 20 + 10 + charge_min
        expected = math.floor(720 / cycle_min)
        self.assertEqual(expected, 21)
        self.assertEqual(mid["trips_per_day"], expected)
        self.assertGreater(mid["trips_per_day"], 15)
        self.assertIsNone(mid["assumptions"]["max_trips_per_day_cap"])
        self.assertFalse(mid["assumptions"]["trips_per_day_capped"])
        self.assertAlmostEqual(mid["assumptions"]["charge_recovery_min"], charge_min, places=3)
        self.assertIn("engineering validation", mid["assumptions"]["charge_recovery_engineering_status"])

    def test_korea_long_route_exact_cycle_formula(self):
        distance = 54.8
        mid = AGGREGATE.run_scenarios(corridor("South Korea", distance), "pioneer_ii")["mid"]
        expected_charge = distance / 70 * 45
        expected_cycle = distance / 20 * 60 + 20 + 10 + expected_charge
        self.assertEqual(mid["trips_per_day"], math.floor(720 / expected_cycle))
        self.assertEqual(mid["trips_per_day"], 3)
        self.assertAlmostEqual(mid["assumptions"]["charge_recovery_min"], expected_charge, places=3)

    def test_korea_midpoint_inputs_are_separate_and_bookends_unchanged(self):
        scenarios = AGGREGATE.run_scenarios(corridor("South Korea", 1.1), "pioneer_ii")
        self.assertEqual(scenarios["mid"]["revenue_inputs"]["load_factor"], 0.65)
        self.assertEqual(scenarios["mid"]["assumptions"]["revenue_leg_pct"], 0.65)
        self.assertEqual(scenarios["thin"]["revenue_inputs"]["load_factor"], 0.45)
        self.assertEqual(scenarios["full"]["revenue_inputs"]["load_factor"], 0.70)
        expected_tpy = round(
            scenarios["mid"]["assumptions"]["gross_legs_per_day"]
            * scenarios["mid"]["assumptions"]["operating_days_yr"]
            * 0.65
        )
        self.assertEqual(scenarios["mid"]["trips_per_year"], expected_tpy)

    def test_non_korean_control_retains_existing_cap_and_midpoint(self):
        scenarios = AGGREGATE.run_scenarios(
            corridor("United Arab Emirates", 1.1), "pioneer_ii"
        )
        mid = scenarios["mid"]
        self.assertEqual(mid["trips_per_day"], 15)
        self.assertEqual(mid["assumptions"]["max_trips_per_day_cap"], 15)
        self.assertTrue(mid["assumptions"]["trips_per_day_capped"])
        self.assertEqual(mid["assumptions"]["charge_recovery_min"], 0.0)
        self.assertEqual(mid["revenue_inputs"]["load_factor"], 0.55)
        self.assertEqual(mid["assumptions"]["revenue_leg_pct"], 0.65)


class TransparentSheetEngineTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            import openpyxl  # noqa: F401
        except ImportError as exc:
            raise unittest.SkipTest("openpyxl is required for transparent Sheet engine test") from exc
        cls.tempdir = tempfile.TemporaryDirectory()
        cls.out = Path(cls.tempdir.name) / "korea-schedule-test.xlsx"
        subprocess.run(
            [
                sys.executable,
                str(FINANCE / "build_transparent_sheet.py"),
                "--partner",
                "korea-schedule-test",
                "--corridors",
                str(FIXTURE),
                "--out",
                str(cls.out),
            ],
            cwd=FINANCE,
            check=True,
        )
        import openpyxl
        cls.wb = openpyxl.load_workbook(cls.out, data_only=False)

    @classmethod
    def tearDownClass(cls):
        if hasattr(cls, "tempdir"):
            cls.tempdir.cleanup()

    def test_visible_headers_keep_schedule_and_utilization_separate(self):
        ws = self.wb["Corridor economics"]
        headers = [cell.value for cell in ws[3]]
        for expected in (
            "Gross legs/day",
            "Revenue-leg utilization",
            "Seat occupancy",
            "Charge recovery min",
        ):
            self.assertIn(expected, headers)

    def test_korea_formulas_and_non_korea_control_are_country_scoped(self):
        ws = self.wb["Corridor economics"]
        headers = {cell.value: cell.column for cell in ws[3]}
        by_corridor = {ws.cell(r, headers["Corridor"]).value: r for r in range(4, ws.max_row + 1)}
        korea_row = by_corridor["Korea Short A → Korea Short B"]
        control_row = by_corridor["Control Short A → Control Short B"]

        def formula(row: int, header: str) -> str:
            return ws.cell(row, headers[header]).value

        for row in (korea_row, control_row):
            self.assertIn('"South Korea"', formula(row, "Gross legs/day"))
            self.assertIn("korea_charge_range_nm", formula(row, "Charge recovery min"))
            self.assertIn("korea_mid_revleg", formula(row, "Revenue-leg utilization"))
            self.assertIn("korea_mid_load", formula(row, "Seat occupancy"))
        self.assertIn("MIN(15", formula(control_row, "Gross legs/day"))
        self.assertIn("FLOOR(60*korea_service_hr", formula(korea_row, "Gross legs/day"))

    def test_rendered_assumptions_call_charging_a_planning_proxy(self):
        ws = self.wb["Assumptions"]
        text = " ".join(str(cell.value or "") for row in ws.iter_rows() for cell in row)
        self.assertIn("PLANNING PROXY", text)
        self.assertIn("engineering validation", text)
        self.assertIn("not a certified charge curve", text)


if __name__ == "__main__":
    unittest.main()
